"""Admin timekeeper routes (employees, jobs, time entries, export, audit)."""

import math
from collections import defaultdict
from io import BytesIO

from flask import (
    Blueprint, abort, flash, jsonify, redirect, render_template, request,
    send_file, url_for,
)
from flask_login import current_user, login_required

import config
import database

time_admin_bp = Blueprint('time_admin', __name__)

# Logo paths for reports
_BDB_LOGO = config.BASE_DIR / "static" / "bdb-logo.png"
_SECTION_COLORS = {
    "green": (22, 163, 74),
    "orange": (234, 88, 12),
    "purple": (124, 58, 237),
    "blue": (37, 99, 235),
}


# ---------------------------------------------------------------------------
# Lazy import helpers to avoid circular imports
# ---------------------------------------------------------------------------

def _helpers():
    import app as _app
    return _app


# ---------------------------------------------------------------------------
# Scheduler role: only allow time-entry and audit-log routes
# ---------------------------------------------------------------------------

_SCHEDULER_ALLOWED = frozenset({
    'time_admin.admin_time_entries',
    'time_admin.admin_time_entry_detail',
    'time_admin.admin_entry_update_notes',
    'time_admin.admin_entry_update_status',
    'time_admin.admin_entry_manual_times',
    'time_admin.admin_manual_entry',
    'time_admin.admin_audit_log',
})


@time_admin_bp.before_request
def _check_scheduler_access():
    """Block scheduler role from employee, job, export, and guide routes."""
    if not current_user.is_authenticated:
        return
    if current_user.is_scheduler and not current_user.is_bdb:
        if request.endpoint not in _SCHEDULER_ALLOWED:
            abort(403)


# ---------------------------------------------------------------------------
# Report helpers (shared by Excel and PDF exports)
# ---------------------------------------------------------------------------

def _company_logo_path(token_str):
    """Return Path to company logo if it exists, else None."""
    td = database.get_token(token_str)
    if td and td.get("logo_file"):
        p = config.LOGOS_DIR / td["logo_file"]
        return p if p.exists() else None
    return None


def _resize_logo(path, max_w=800, max_h=800, target_kb=750):
    """Resize a logo for reports. Keeps it sharp but targets ~750KB max."""
    from PIL import Image as PILImage
    from io import BytesIO
    img = PILImage.open(str(path))
    # Flatten RGBA to RGB if alpha is fully opaque (or paste onto white)
    if img.mode == "RGBA":
        alpha = img.getchannel("A")
        if alpha.getextrema() == (255, 255):
            img = img.convert("RGB")
        else:
            flat = PILImage.new("RGB", img.size, (255, 255, 255))
            flat.paste(img, mask=img.split()[3])
            img = flat
    else:
        img = img.convert("RGB")
    img.thumbnail((max_w, max_h), PILImage.LANCZOS)
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=95, optimize=True)
    if buf.getbuffer().nbytes > target_kb * 1024:
        for q in (85, 75, 65):
            buf = BytesIO()
            img.save(buf, format="JPEG", quality=q, optimize=True)
            if buf.getbuffer().nbytes <= target_kb * 1024:
                break
    buf.seek(0)
    buf.name = "logo.jpg"
    return buf, img.size[0], img.size[1]


def _safe(text):
    """Replace non-latin-1 chars so Helvetica won't choke."""
    return str(text).encode("latin-1", "replace").decode("latin-1")


def _xl_add_logos(ws, token_str, last_row, logo_col="H"):
    """Add company logo (top-right) and BDB logo (bottom-center) to Excel sheet."""
    from openpyxl.drawing.image import Image as XlImage
    logo = _company_logo_path(token_str)
    if logo:
        buf, w, h = _resize_logo(logo)
        img = XlImage(buf)
        # Display small but keep high-res source for clarity
        scale = min(120 / w, 50 / h)
        img.width = int(w * scale)
        img.height = int(h * scale)
        ws.add_image(img, f"{logo_col}1")
    if _BDB_LOGO.exists():
        buf, w, h = _resize_logo(_BDB_LOGO)
        bdb = XlImage(buf)
        scale = min(200 / w, 60 / h)
        bdb.width = int(w * scale)
        bdb.height = int(h * scale)
        ws.add_image(bdb, f"C{last_row + 2}")


def _pdf_add_logos(pdf, company_logo):
    """Add company logo top-right on current page (does not move cursor)."""
    if company_logo and company_logo.exists():
        buf, w, h = _resize_logo(company_logo)
        logo_w = 30
        saved_y = pdf.get_y()
        pdf.image(buf, x=pdf.w - pdf.r_margin - logo_w, y=pdf.t_margin, w=logo_w)
        pdf.set_y(saved_y)


def _pdf_add_bdb_footer(pdf):
    """Add BDB logo centered at the bottom of the current page."""
    if _BDB_LOGO.exists():
        buf, w, h = _resize_logo(_BDB_LOGO)
        x = (pdf.w - 50) / 2
        pdf.image(buf, x=x, y=pdf.h - 22, w=50)


def _pdf_section_header(pdf, title, color_rgb):
    """Draw a colored section title bar. Adds a page if not enough room."""
    # Need ~30mm for section header + table header + at least one data row
    if pdf.get_y() + 30 > pdf.h - pdf.b_margin:
        pdf.add_page()
    pdf.ln(4)
    pdf.set_fill_color(*color_rgb)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, f"  {_safe(title)}", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)


def _pdf_table_header(pdf, headers, widths, color_rgb):
    """Draw a table header row with colored fill."""
    pdf.set_fill_color(*color_rgb)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 7, h, border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)


# ---------------------------------------------------------------------------
# GPS distance calculation
# ---------------------------------------------------------------------------

def _haversine_miles(lat1, lng1, lat2, lng2):
    if any(v is None for v in (lat1, lng1, lat2, lng2)):
        return None
    R = 3958.8  # Earth radius in miles
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlng / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ---------------------------------------------------------------------------
# Employee Management
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/employees")
@login_required
def admin_employees():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    token_str, selected_token = _app._get_selected_token(tokens)
    employees = database.get_employees_by_token(token_str) if token_str else []
    return render_template(
        "admin/employees.html",
        tokens=tokens, selected_token=selected_token, employees=employees,
    )


@time_admin_bp.route("/admin/employees/create", methods=["POST"])
@login_required
def admin_employee_create():
    _app = _helpers()
    token_str = request.form.get("token", "").strip()
    _app._verify_token_access(token_str)
    name = request.form.get("name", "").strip()
    employee_id_str = request.form.get("employee_id", "").strip()
    username = request.form.get("username", "").strip() or None
    password = request.form.get("password", "").strip() or None
    wage_str = request.form.get("hourly_wage", "").strip()
    hourly_wage = float(wage_str) if wage_str else None
    if not name or not employee_id_str or not token_str:
        flash("Name, Employee ID, and company are required.", "error")
    elif username and not password:
        flash("Password is required when setting a username.", "error")
    elif username and not database.check_employee_username_available(username, token_str):
        flash(f"Username '{username}' is already taken.", "error")
    else:
        receipt_access = 1 if request.form.get("receipt_access") else 0
        timekeeper_access = 1 if request.form.get("timekeeper_access") else 0
        job_photos_access = 1 if request.form.get("job_photos_access") else 0
        schedule_access = 1 if request.form.get("schedule_access") else 0
        estimate_access = 1 if request.form.get("estimate_access") else 0
        database.create_employee(name, employee_id_str, token_str, username, password,
                                 hourly_wage=hourly_wage, receipt_access=receipt_access,
                                 timekeeper_access=timekeeper_access,
                                 job_photos_access=job_photos_access,
                                 schedule_access=schedule_access,
                                 estimate_access=estimate_access)
        flash(f"Employee {name} added.", "success")
    return redirect(url_for("time_admin.admin_employees", token=token_str))


@time_admin_bp.route("/admin/employees/<int:emp_id>/update", methods=["POST"])
@login_required
def admin_employee_update(emp_id):
    _app = _helpers()
    emp = database.get_employee(emp_id)
    if emp:
        _app._verify_token_access(emp["token"])
    kwargs = {}
    toggle_fields = ("receipt_access", "timekeeper_access", "job_photos_access",
                      "schedule_access", "estimate_access")
    if request.is_json:
        data = request.get_json()
        name = data.get("name", "").strip()
        employee_id_str = data.get("employee_id", "").strip()
        if "hourly_wage" in data:
            wage_raw = data["hourly_wage"]
            kwargs["hourly_wage"] = float(wage_raw) if wage_raw is not None and str(wage_raw).strip() != "" else None
        for field in toggle_fields:
            if field in data:
                kwargs[field] = int(data[field])
    else:
        name = request.form.get("name", "").strip()
        employee_id_str = request.form.get("employee_id", "").strip()
        if "hourly_wage" in request.form:
            wage_str = request.form["hourly_wage"].strip()
            kwargs["hourly_wage"] = float(wage_str) if wage_str else None
        for field in toggle_fields:
            if field in request.form:
                kwargs[field] = int(request.form[field])
    database.update_employee(emp_id, name, employee_id_str, **kwargs)
    if request.is_json:
        return jsonify({"success": True})
    flash("Employee updated.", "success")
    return redirect(url_for("time_admin.admin_employees", token=emp["token"] if emp else ""))


@time_admin_bp.route("/admin/employees/<int:emp_id>/toggle", methods=["POST"])
@login_required
def admin_employee_toggle(emp_id):
    _app = _helpers()
    emp = database.get_employee(emp_id)
    if emp:
        _app._verify_token_access(emp["token"])
    database.toggle_employee(emp_id)
    flash("Employee status toggled.", "success")
    return redirect(url_for("time_admin.admin_employees", token=emp["token"] if emp else ""))


@time_admin_bp.route("/admin/employees/<int:emp_id>/set-credentials", methods=["POST"])
@login_required
def admin_employee_set_credentials(emp_id):
    _app = _helpers()
    emp = database.get_employee(emp_id)
    if not emp:
        abort(404)
    _app._verify_token_access(emp["token"])
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    if not username or not password:
        flash("Username and password are required.", "error")
    elif not database.check_employee_username_available(username, emp["token"], exclude_emp_id=emp_id):
        flash(f"Username '{username}' is already taken.", "error")
    else:
        database.set_employee_credentials(emp_id, username, password)
        flash(f"Login credentials set for {emp['name']}.", "success")
    return redirect(url_for("time_admin.admin_employees", token=emp["token"]))


@time_admin_bp.route("/admin/employees/<int:emp_id>/reset-password", methods=["POST"])
@login_required
def admin_employee_reset_password(emp_id):
    _app = _helpers()
    emp = database.get_employee(emp_id)
    if not emp:
        abort(404)
    _app._verify_token_access(emp["token"])
    new_password = request.form.get("new_password", "").strip()
    if not new_password:
        flash("New password is required.", "error")
    else:
        database.reset_employee_password(emp_id, new_password)
        flash(f"Password reset for {emp['name']}.", "success")
    return redirect(url_for("time_admin.admin_employees", token=emp["token"]))


# ---------------------------------------------------------------------------
# Job Management
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/jobs")
@login_required
def admin_jobs():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    token_str, selected_token = _app._get_selected_token(tokens)
    jobs = database.get_jobs_with_customer(token_str) if token_str else []
    customers = database.get_customers_by_token(token_str, active_only=True) if token_str else []
    return render_template(
        "admin/jobs.html",
        tokens=tokens, selected_token=selected_token, jobs=jobs, customers=customers,
    )


@time_admin_bp.route("/admin/jobs/create", methods=["POST"])
@login_required
def admin_job_create():
    _app = _helpers()
    token_str = request.form.get("token", "").strip()
    _app._verify_token_access(token_str)
    job_name = request.form.get("job_name", "").strip()
    job_address = request.form.get("job_address", "").strip()
    lat = request.form.get("latitude", "").strip()
    lng = request.form.get("longitude", "").strip()

    if not job_name or not job_address or not token_str:
        flash("Job name, address, and token are required.", "error")
        return redirect(url_for("time_admin.admin_jobs", token=token_str))

    latitude = float(lat) if lat else None
    longitude = float(lng) if lng else None
    customer_id = request.form.get("customer_id", type=int)
    database.create_job(job_name, job_address, latitude, longitude, token_str, customer_id=customer_id)
    flash(f"Job '{job_name}' created.", "success")
    return redirect(url_for("time_admin.admin_jobs", token=token_str))


@time_admin_bp.route("/admin/jobs/<int:job_id>/update", methods=["POST"])
@login_required
def admin_job_update(job_id):
    _app = _helpers()
    job_check = database.get_job(job_id)
    if job_check:
        _app._verify_token_access(job_check["token"])
    if request.is_json:
        data = request.get_json()
        job_name = data.get("job_name", "").strip()
        job_address = data.get("job_address", "").strip()
        lat = data.get("latitude")
        lng = data.get("longitude")
        customer_id = data.get("customer_id") if data.get("customer_id") else None
    else:
        job_name = request.form.get("job_name", "").strip()
        job_address = request.form.get("job_address", "").strip()
        lat = request.form.get("latitude", "").strip()
        lng = request.form.get("longitude", "").strip()
        customer_id = request.form.get("customer_id", type=int)
    latitude = float(lat) if lat else None
    longitude = float(lng) if lng else None
    database.update_job(job_id, job_name, job_address, latitude, longitude, customer_id=customer_id)
    if request.is_json:
        return jsonify({"success": True})
    flash("Job updated.", "success")
    job = database.get_job(job_id)
    return redirect(url_for("time_admin.admin_jobs", token=job["token"] if job else ""))


@time_admin_bp.route("/admin/jobs/<int:job_id>/toggle", methods=["POST"])
@login_required
def admin_job_toggle(job_id):
    _app = _helpers()
    job = database.get_job(job_id)
    if job:
        _app._verify_token_access(job["token"])
    database.toggle_job(job_id)
    flash("Job status toggled.", "success")
    return redirect(url_for("time_admin.admin_jobs", token=job["token"] if job else ""))


@time_admin_bp.route("/admin/jobs/<int:job_id>/archive", methods=["POST"])
@login_required
def admin_job_archive(job_id):
    _app = _helpers()
    job = database.get_job(job_id)
    if job:
        _app._verify_token_access(job["token"])
    database.archive_job(job_id)
    flash("Job archived.", "success")
    return redirect(url_for("time_admin.admin_jobs", token=job["token"] if job else ""))


@time_admin_bp.route("/admin/jobs/<int:job_id>/unarchive", methods=["POST"])
@login_required
def admin_job_unarchive(job_id):
    _app = _helpers()
    job = database.get_job(job_id)
    if job:
        _app._verify_token_access(job["token"])
    database.unarchive_job(job_id)
    flash("Job unarchived.", "success")
    return redirect(url_for("time_admin.admin_jobs", token=job["token"] if job else ""))


# ---------------------------------------------------------------------------
# Time Entries - Admin Browse
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/time-entries")
@login_required
def admin_time_entries():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    token_str, selected_token = _app._get_selected_token(tokens)

    employee_id = request.args.get("employee_id", type=int)
    job_id = request.args.get("job_id", type=int)
    status = request.args.get("status", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    entries = []
    employees = []
    jobs = []
    if token_str:
        entries = database.get_time_entries(
            token_str, employee_id=employee_id, job_id=job_id,
            status=status or None, date_from=date_from or None,
            date_to=date_to or None, limit=200,
        )
        employees = database.get_employees_by_token(token_str)
        jobs = database.get_jobs_by_token(token_str)

    # Calculate GPS distances for entries
    for entry in entries:
        entry.setdefault("clock_in_distance", None)
        entry.setdefault("clock_in_flagged", False)
        if entry.get("clock_in_lat") and entry.get("job_id"):
            job = database.get_job(entry["job_id"])
            if job and job.get("latitude"):
                dist = _haversine_miles(
                    entry["clock_in_lat"], entry["clock_in_lng"],
                    job["latitude"], job["longitude"],
                )
                entry["clock_in_distance"] = round(dist, 2) if dist is not None else None
                entry["clock_in_flagged"] = (
                    dist is not None and dist > config.GPS_FLAG_DISTANCE_MILES
                )

    filters = {
        "employee_id": str(employee_id) if employee_id else "",
        "job_id": str(job_id) if job_id else "",
        "status": status,
        "date_from": date_from,
        "date_to": date_to,
    }
    return render_template(
        "admin/time_entries.html",
        tokens=tokens, selected_token=selected_token,
        entries=entries, employees=employees, jobs=jobs,
        filters=filters,
    )


# ---------------------------------------------------------------------------
# Time Entry Detail
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/time-entries/<int:entry_id>")
@login_required
def admin_time_entry_detail(entry_id):
    _app = _helpers()
    entry = database.get_time_entry(entry_id)
    if not entry:
        abort(404)
    _app._verify_token_access(entry["token"])

    # Calculate distances (always set keys so template doesn't get UndefinedError)
    entry.setdefault("clock_in_distance", None)
    entry.setdefault("clock_out_distance", None)
    if entry.get("clock_in_lat") and entry.get("job_lat"):
        entry["clock_in_distance"] = _haversine_miles(
            entry["clock_in_lat"], entry["clock_in_lng"],
            entry["job_lat"], entry["job_lng"],
        )
    if entry.get("clock_out_lat") and entry.get("job_lat"):
        entry["clock_out_distance"] = _haversine_miles(
            entry["clock_out_lat"], entry["clock_out_lng"],
            entry["job_lat"], entry["job_lng"],
        )

    audit = database.get_audit_log(entry["token"])
    audit = [a for a in audit if a.get("time_entry_id") == entry_id]

    jobs = database.get_jobs_by_token(entry["token"])

    tokens = _app._get_tokens_for_user()
    _, selected_token = _app._get_selected_token(tokens)

    return render_template(
        "admin/time_entry_detail.html", entry=entry, audit=audit,
        jobs=jobs, tokens=tokens, selected_token=selected_token,
    )


# ---------------------------------------------------------------------------
# Time Entry Actions
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/time-entries/<int:entry_id>/delete", methods=["POST"])
@login_required
def admin_entry_delete(entry_id):
    _app = _helpers()
    entry = database.get_time_entry(entry_id)
    if not entry:
        abort(404)
    _app._verify_token_access(entry["token"])
    if not current_user.is_admin and not current_user.is_bdb:
        abort(403)
    reason = request.form.get("reason", "").strip()
    if not reason:
        flash("A reason is required to delete a time entry.", "error")
        return redirect(url_for("time_admin.admin_time_entry_detail", entry_id=entry_id))
    database.delete_time_entry(entry_id, entry["token"], current_user.username, reason)
    flash(f"Time entry #{entry_id} deleted.", "success")
    return redirect(url_for("time_admin.admin_time_entries", token=entry["token"]))


@time_admin_bp.route("/admin/time-entries/<int:entry_id>/notes", methods=["POST"])
@login_required
def admin_entry_update_notes(entry_id):
    _app = _helpers()
    entry = database.get_time_entry(entry_id)
    if entry:
        _app._verify_token_access(entry["token"])
    notes = request.form.get("admin_notes", "").strip()
    reason = request.form.get("reason", "").strip()
    database.update_entry_notes(entry_id, notes, current_user.username, reason)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"success": True})
    flash("Notes updated.", "success")
    return redirect(url_for("time_admin.admin_time_entry_detail", entry_id=entry_id))


@time_admin_bp.route("/admin/time-entries/<int:entry_id>/status", methods=["POST"])
@login_required
def admin_entry_update_status(entry_id):
    _app = _helpers()
    entry = database.get_time_entry(entry_id)
    if entry:
        _app._verify_token_access(entry["token"])
    new_status = request.form.get("status", "").strip()
    reason = request.form.get("reason", "").strip()
    if new_status not in ("active", "completed", "needs_review"):
        flash("Invalid status.", "error")
    elif not reason:
        flash("A reason is required to change status.", "error")
    else:
        database.update_entry_status(entry_id, new_status, current_user.username, reason)
        flash(f"Status changed to {new_status}.", "success")
    return redirect(url_for("time_admin.admin_time_entry_detail", entry_id=entry_id))


@time_admin_bp.route("/admin/time-entries/<int:entry_id>/change-job", methods=["POST"])
@login_required
def admin_entry_change_job(entry_id):
    _app = _helpers()
    entry = database.get_time_entry(entry_id)
    if not entry:
        abort(404)
    _app._verify_token_access(entry["token"])
    new_job_id = request.form.get("job_id", "").strip()
    reason = request.form.get("reason", "").strip()
    if not new_job_id:
        flash("Please select a job.", "error")
    elif not reason:
        flash("A reason is required to change the job.", "error")
    else:
        try:
            new_job_id = int(new_job_id)
        except ValueError:
            flash("Invalid job.", "error")
            return redirect(url_for("time_admin.admin_time_entry_detail", entry_id=entry_id))
        if new_job_id == entry["job_id"]:
            flash("Selected job is the same as the current job.", "error")
        else:
            job = database.get_job(new_job_id)
            if not job:
                flash("Job not found.", "error")
            else:
                database.update_entry_job(entry_id, new_job_id, current_user.username, reason)
                flash(f"Job changed to {job['job_name']}.", "success")
    return redirect(url_for("time_admin.admin_time_entry_detail", entry_id=entry_id))


@time_admin_bp.route("/admin/time-entries/<int:entry_id>/manual-times", methods=["POST"])
@login_required
def admin_entry_manual_times(entry_id):
    _app = _helpers()
    entry = database.get_time_entry(entry_id)
    if entry:
        _app._verify_token_access(entry["token"])
    manual_in = request.form.get("manual_time_in", "").strip()
    manual_out = request.form.get("manual_time_out", "").strip()
    notes = request.form.get("admin_notes", "").strip()
    if not manual_in and not manual_out:
        flash("At least one time is required.", "error")
    elif not notes:
        flash("A reason is required.", "error")
    else:
        database.update_manual_times(entry_id, manual_in, manual_out, notes, current_user.username, reason=notes)
        changed = []
        if manual_in:
            changed.append("Clock In")
        if manual_out:
            changed.append("Clock Out")
        flash(f"Manual {' & '.join(changed)} updated.", "success")
    return redirect(url_for("time_admin.admin_time_entry_detail", entry_id=entry_id))


# ---------------------------------------------------------------------------
# Manual Entry (Admin creates)
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/time-entries/manual", methods=["GET", "POST"])
@login_required
def admin_manual_entry():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    if not current_user.is_bdb:
        token_str = current_user.token
    else:
        token_str = request.args.get("token", "") or request.form.get("token", "")
        if not token_str and tokens:
            token_str = tokens[0]["token"]
    selected_token = database.get_token(token_str) if token_str else None

    employees = database.get_employees_by_token(token_str, active_only=True) if token_str else []
    jobs = database.get_jobs_by_token(token_str, active_only=True) if token_str else []

    if request.method == "POST":
        _app._verify_token_access(token_str)
        employee_id = request.form.get("employee_id", type=int)
        job_id = request.form.get("job_id", type=int)
        manual_in = request.form.get("manual_time_in", "").strip()
        manual_out = request.form.get("manual_time_out", "").strip()
        notes = request.form.get("admin_notes", "").strip()

        errors = []
        if not token_str:
            errors.append("Company token missing — navigate back and reload the page.")
        if not employee_id:
            errors.append("Employee is required.")
        if not job_id:
            errors.append("Job is required.")
        if not manual_in:
            errors.append("Clock-in time is required.")
        if not notes:
            errors.append("Admin notes are required.")
        if errors:
            for e in errors:
                flash(e, "error")
        else:
            entry_id = database.create_manual_entry(
                employee_id, job_id, token_str, manual_in, manual_out or None,
                notes, current_user.username, reason=notes,
            )
            if manual_out:
                flash("Manual entry created (clocked in and out).", "success")
            else:
                flash("Clock-in entry created. Employee is now shown as active.", "success")
            return redirect(url_for("time_admin.admin_time_entries", token=token_str))

    return render_template(
        "admin/manual_entry.html",
        tokens=tokens, selected_token=selected_token,
        employees=employees, jobs=jobs,
    )


# ---------------------------------------------------------------------------
# Audit Log
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/audit-log")
@login_required
def admin_audit_log():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    token_str, selected_token = _app._get_selected_token(tokens)
    logs = database.get_audit_log(token_str) if token_str else []
    return render_template(
        "admin/audit_log.html",
        tokens=tokens, selected_token=selected_token, logs=logs,
    )


# ---------------------------------------------------------------------------
# Export Page
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export")
@login_required
def admin_export():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    token_str, selected_token = _app._get_selected_token(tokens)
    return render_template(
        "admin/export.html",
        tokens=tokens, selected_token=selected_token,
    )


# ---------------------------------------------------------------------------
# Excel Export Download
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export/download")
@login_required
def admin_export_download():
    _app = _helpers()
    token_str = request.args.get("token", "")
    if not current_user.is_bdb:
        token_str = current_user.token
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    _app._verify_token_access(token_str)

    if not token_str:
        flash("Token is required.", "error")
        return redirect(url_for("time_admin.admin_export"))

    entries = database.get_time_entries_for_export(
        token_str, date_from=date_from or None, date_to=date_to or None,
    )
    token_data = database.get_token(token_str)
    company = token_data["company_name"] if token_data else "Unknown"

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.views import SheetView

    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Employee Name", "Employee ID", "Job Name", "Job Address",
        "Clock In", "Clock In Method", "Clock In GPS",
        "Clock Out", "Clock Out Method", "Clock Out GPS",
        "Manual In", "Manual Out",
        "Total Hours", "Status", "Admin Notes",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, e in enumerate(entries, 2):
        clock_in_gps = ""
        if e.get("clock_in_lat"):
            clock_in_gps = f"{e['clock_in_lat']}, {e['clock_in_lng']}"
        clock_out_gps = ""
        if e.get("clock_out_lat"):
            clock_out_gps = f"{e['clock_out_lat']}, {e['clock_out_lng']}"

        row_data = [
            e.get("employee_name", ""),
            e.get("emp_id_str", ""),
            e.get("job_name", ""),
            e.get("job_address", ""),
            e.get("clock_in_time", ""),
            e.get("clock_in_method", ""),
            clock_in_gps,
            e.get("clock_out_time", ""),
            e.get("clock_out_method", ""),
            clock_out_gps,
            e.get("manual_time_in", ""),
            e.get("manual_time_out", ""),
            e.get("total_hours"),
            e.get("status", ""),
            e.get("admin_notes", ""),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # --- Employee Summary Section ---
    summary_start = len(entries) + 4  # 2 blank rows after data
    summary_header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    bold_font = Font(bold=True, size=11)

    ws.cell(row=summary_start - 1, column=1, value="Employee Summary").font = Font(bold=True, size=13)

    sum_headers = ["Employee Name", "Employee ID", "Total Hours"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws.cell(row=summary_start, column=col, value=h)
        cell.font = summary_header_font
        cell.fill = summary_fill
        cell.alignment = header_align
        cell.border = thin_border

    emp_totals = defaultdict(lambda: {"name": "", "emp_id": "", "hours": 0.0})
    for e in entries:
        key = e.get("emp_id_str", "")
        emp_totals[key]["name"] = e.get("employee_name", "")
        emp_totals[key]["emp_id"] = key
        emp_totals[key]["hours"] += float(e.get("total_hours") or 0)

    sorted_emps = sorted(emp_totals.values(), key=lambda x: x["name"].lower())
    company_total = 0.0
    for i, emp in enumerate(sorted_emps):
        r = summary_start + 1 + i
        ws.cell(row=r, column=1, value=emp["name"]).border = thin_border
        ws.cell(row=r, column=2, value=emp["emp_id"]).border = thin_border
        hrs_cell = ws.cell(row=r, column=3, value=round(emp["hours"], 2))
        hrs_cell.border = thin_border
        company_total += emp["hours"]

    total_row = summary_start + 1 + len(sorted_emps)
    ws.cell(row=total_row, column=1, value="Company Total").font = bold_font
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=2).border = thin_border
    total_cell = ws.cell(row=total_row, column=3, value=round(company_total, 2))
    total_cell.font = bold_font
    total_cell.border = thin_border

    # --- Employee Hours by Job Section ---
    emp_job_start = total_row + 3  # 2 blank rows after Company Total
    emp_job_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")

    ws.cell(row=emp_job_start - 1, column=1, value="Employee Hours by Job").font = Font(bold=True, size=13)

    ej_headers = ["Employee Name", "Employee ID", "Job Name", "Hours"]
    for col, h in enumerate(ej_headers, 1):
        cell = ws.cell(row=emp_job_start, column=col, value=h)
        cell.font = summary_header_font
        cell.fill = emp_job_fill
        cell.alignment = header_align
        cell.border = thin_border

    emp_job_totals = defaultdict(lambda: {"name": "", "emp_id": "", "job": "", "hours": 0.0})
    for e in entries:
        key = (e.get("emp_id_str", ""), e.get("job_name", ""))
        emp_job_totals[key]["name"] = e.get("employee_name", "")
        emp_job_totals[key]["emp_id"] = e.get("emp_id_str", "")
        emp_job_totals[key]["job"] = e.get("job_name", "")
        emp_job_totals[key]["hours"] += float(e.get("total_hours") or 0)

    sorted_emp_jobs = sorted(emp_job_totals.values(), key=lambda x: (x["name"].lower(), x["job"].lower()))
    for i, ej in enumerate(sorted_emp_jobs):
        r = emp_job_start + 1 + i
        ws.cell(row=r, column=1, value=ej["name"]).border = thin_border
        ws.cell(row=r, column=2, value=ej["emp_id"]).border = thin_border
        ws.cell(row=r, column=3, value=ej["job"]).border = thin_border
        hrs_cell = ws.cell(row=r, column=4, value=round(ej["hours"], 2))
        hrs_cell.border = thin_border

    # --- Company Hours by Job Section ---
    cj_start = emp_job_start + 1 + len(sorted_emp_jobs) + 2  # 2 blank rows
    cj_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

    ws.cell(row=cj_start - 1, column=1, value="Company Hours by Job").font = Font(bold=True, size=13)

    cj_headers = ["Job Name", "Total Hours"]
    for col, h in enumerate(cj_headers, 1):
        cell = ws.cell(row=cj_start, column=col, value=h)
        cell.font = summary_header_font
        cell.fill = cj_fill
        cell.alignment = header_align
        cell.border = thin_border

    job_totals = defaultdict(float)
    for e in entries:
        job_totals[e.get("job_name", "")] += float(e.get("total_hours") or 0)

    sorted_jobs = sorted(job_totals.items(), key=lambda x: x[0].lower())
    for i, (job_name, hours) in enumerate(sorted_jobs):
        r = cj_start + 1 + i
        ws.cell(row=r, column=1, value=job_name).border = thin_border
        hrs_cell = ws.cell(row=r, column=2, value=round(hours, 2))
        hrs_cell.border = thin_border

    cj_total_row = cj_start + 1 + len(sorted_jobs)
    ws.cell(row=cj_total_row, column=1, value="Company Total").font = bold_font
    ws.cell(row=cj_total_row, column=1).border = thin_border
    cj_total_cell = ws.cell(row=cj_total_row, column=2, value=round(company_total, 2))
    cj_total_cell.font = bold_font
    cj_total_cell.border = thin_border

    # Add logos
    _xl_add_logos(ws, token_str, cj_total_row, logo_col="L")

    # Set 140% zoom on ALL worksheets
    for sheet in wb.worksheets:
        sheet.sheet_view.zoomScale = 140

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_range = ""
    if date_from:
        date_range += f"_{date_from}"
    if date_to:
        date_range += f"_to_{date_to}"
    filename = f"timekeeper_{company.replace(' ', '_')}{date_range}.xlsx"

    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# PDF Export Download
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export/download-pdf")
@login_required
def admin_export_download_pdf():
    _app = _helpers()
    token_str = request.args.get("token", "")
    if not current_user.is_bdb:
        token_str = current_user.token
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    _app._verify_token_access(token_str)

    if not token_str:
        flash("Token is required.", "error")
        return redirect(url_for("time_admin.admin_export"))

    entries = database.get_time_entries_for_export(
        token_str, date_from=date_from or None, date_to=date_to or None,
    )
    token_data = database.get_token(token_str)
    company = token_data["company_name"] if token_data else "Unknown"
    company_logo = _company_logo_path(token_str)

    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=28)
    pdf.add_page()

    # Title / header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _safe(f"{company} - Hours Report"), ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    range_label = ""
    if date_from:
        range_label += date_from
    if date_to:
        range_label += f"  to  {date_to}"
    if range_label:
        pdf.cell(0, 6, range_label, ln=True, align="C")
    pdf.ln(4)

    # Time entries table (blue header)
    col_widths = [38, 25, 40, 34, 34, 18, 22, 59]
    te_headers = ["Employee", "Emp ID", "Job", "Clock In", "Clock Out",
                  "Hours", "Status", "Notes"]
    _pdf_table_header(pdf, te_headers, col_widths, _SECTION_COLORS["blue"])

    def _fmt_dt(val):
        """Format ISO datetime to 'MM/DD h:MM AM/PM'."""
        s = str(val or "")
        if not s or s == "None":
            return "\u2014"
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(s)
            return dt.strftime("%m/%d %-I:%M %p")
        except Exception:
            return s[:16]

    pdf.set_font("Helvetica", "", 7)
    for e in entries:
        hours_val = float(e.get("total_hours") or 0)
        row = [
            str(e.get("employee_name", ""))[:24],
            str(e.get("emp_id_str", "")),
            str(e.get("job_name", ""))[:24],
            _fmt_dt(e.get("clock_in_time")),
            _fmt_dt(e.get("clock_out_time")),
            f"{hours_val:.2f}",
            str(e.get("status", "")),
            str(e.get("admin_notes", "") or ""),
        ]
        for i, val in enumerate(row):
            pdf.cell(col_widths[i], 6, _safe(val), border=1)
        pdf.ln()

    # Employee summary (green header)
    _pdf_section_header(pdf, "Employee Summary", _SECTION_COLORS["green"])
    sum_widths = [70, 50, 40]
    _pdf_table_header(pdf, ["Employee Name", "Employee ID", "Total Hours"], sum_widths, _SECTION_COLORS["green"])

    emp_totals = defaultdict(lambda: {"name": "", "emp_id": "", "hours": 0.0})
    for e in entries:
        key = e.get("emp_id_str", "")
        emp_totals[key]["name"] = e.get("employee_name", "")
        emp_totals[key]["emp_id"] = key
        emp_totals[key]["hours"] += float(e.get("total_hours") or 0)

    sorted_emps = sorted(emp_totals.values(), key=lambda x: x["name"].lower())
    company_total = 0.0
    pdf.set_font("Helvetica", "", 9)
    for emp in sorted_emps:
        pdf.cell(sum_widths[0], 6, _safe(emp["name"]), border=1)
        pdf.cell(sum_widths[1], 6, _safe(emp["emp_id"]), border=1)
        pdf.cell(sum_widths[2], 6, f"{emp['hours']:.2f}", border=1, align="R")
        pdf.ln()
        company_total += emp["hours"]

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(sum_widths[0] + sum_widths[1], 7, "Company Total", border=1)
    pdf.cell(sum_widths[2], 7, f"{company_total:.2f}", border=1, align="R")
    pdf.ln()

    # Employee Hours by Job (orange header)
    _pdf_section_header(pdf, "Employee Hours by Job", _SECTION_COLORS["orange"])
    ej_widths = [60, 40, 60, 30]
    _pdf_table_header(pdf, ["Employee Name", "Emp ID", "Job", "Hours"], ej_widths, _SECTION_COLORS["orange"])

    emp_job_totals = defaultdict(lambda: {"name": "", "emp_id": "", "job": "", "hours": 0.0})
    for e in entries:
        key = (e.get("emp_id_str", ""), e.get("job_name", ""))
        emp_job_totals[key]["name"] = e.get("employee_name", "")
        emp_job_totals[key]["emp_id"] = e.get("emp_id_str", "")
        emp_job_totals[key]["job"] = e.get("job_name", "")
        emp_job_totals[key]["hours"] += float(e.get("total_hours") or 0)

    sorted_emp_jobs = sorted(emp_job_totals.values(), key=lambda x: (x["name"].lower(), x["job"].lower()))
    pdf.set_font("Helvetica", "", 9)
    for ej in sorted_emp_jobs:
        pdf.cell(ej_widths[0], 6, _safe(ej["name"]), border=1)
        pdf.cell(ej_widths[1], 6, _safe(ej["emp_id"]), border=1)
        pdf.cell(ej_widths[2], 6, _safe(ej["job"]), border=1)
        pdf.cell(ej_widths[3], 6, f"{ej['hours']:.2f}", border=1, align="R")
        pdf.ln()

    # Company Hours by Job (purple header)
    _pdf_section_header(pdf, "Company Hours by Job", _SECTION_COLORS["purple"])
    cj_widths = [100, 40]
    _pdf_table_header(pdf, ["Job Name", "Total Hours"], cj_widths, _SECTION_COLORS["purple"])

    job_totals = defaultdict(float)
    for e in entries:
        job_totals[e.get("job_name", "")] += float(e.get("total_hours") or 0)

    sorted_jobs = sorted(job_totals.items(), key=lambda x: x[0].lower())
    pdf.set_font("Helvetica", "", 9)
    for job_name, hours in sorted_jobs:
        pdf.cell(cj_widths[0], 6, _safe(job_name), border=1)
        pdf.cell(cj_widths[1], 6, f"{hours:.2f}", border=1, align="R")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(cj_widths[0], 7, "Company Total", border=1)
    pdf.cell(cj_widths[1], 7, f"{company_total:.2f}", border=1, align="R")
    pdf.ln()

    # Company logo on page 1 (overlay — placed last so it doesn't affect layout)
    current_page = pdf.page
    pdf.page = 1
    saved_y = pdf.get_y()
    if company_logo and company_logo.exists():
        buf, w, h = _resize_logo(company_logo)
        pdf.image(buf, x=pdf.w - 30, y=5, w=20)
    pdf.page = current_page
    pdf.set_y(saved_y)

    # BDB logo footer
    _pdf_add_bdb_footer(pdf)

    output = BytesIO()
    pdf.output(output)
    output.seek(0)

    date_range = ""
    if date_from:
        date_range += f"_{date_from}"
    if date_to:
        date_range += f"_to_{date_to}"
    filename = f"timekeeper_{company.replace(' ', '_')}{date_range}.pdf"

    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/pdf",
    )


# ---------------------------------------------------------------------------
# Payroll Cost Report Export
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export/payroll-cost")
@login_required
def admin_export_payroll_cost():
    _app = _helpers()
    token_str = request.args.get("token", "")
    if not current_user.is_bdb:
        token_str = current_user.token
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    _app._verify_token_access(token_str)

    if not token_str:
        flash("Token is required.", "error")
        return redirect(url_for("time_admin.admin_export"))

    entries = database.get_time_entries_for_export(
        token_str, date_from=date_from or None, date_to=date_to or None,
    )
    token_data = database.get_token(token_str)
    company = token_data["company_name"] if token_data else "Unknown"
    burden_pct = token_data.get("labor_burden_pct", 0) if token_data else 0

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Cost Report"

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True, size=11)
    money_fmt = '#,##0.00'

    # Title rows
    ws.cell(row=1, column=1, value=f"PAYROLL COST ESTIMATE — {company}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="NOT FOR BOOKKEEPING PURPOSES — Estimate Only").font = Font(bold=True, size=11, color="DC2626")
    range_label = ""
    if date_from:
        range_label += date_from
    if date_to:
        range_label += f" to {date_to}"
    ws.cell(row=3, column=1, value=f"Date range: {range_label}").font = Font(size=10, color="6B7280")
    ws.cell(row=4, column=1, value=f"Labor burden: {burden_pct}%").font = Font(size=10, color="6B7280")

    # Pre-compute OT effective rates for all entries
    eff_rates = database.get_effective_rates_for_entries(token_str, entries)

    def _entry_base_pay(e):
        """Return OT-adjusted base pay for a single time entry."""
        hrs = float(e.get("total_hours") or 0)
        if hrs <= 0 or e.get("hourly_wage") is None:
            return 0.0
        week = database._get_week_start_sunday(e["clock_in_time"])
        rate_info = eff_rates.get((e["employee_id"], week))
        if rate_info and rate_info["effective_rate"]:
            return hrs * rate_info["effective_rate"]
        return hrs * e["hourly_wage"]

    # --- Section 1: Employee Cost Summary ---
    s1_start = 6
    s1_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    s1_font = Font(bold=True, color="FFFFFF", size=11)
    ws.cell(row=s1_start, column=1, value="Employee Cost Summary").font = Font(bold=True, size=13)
    s1_start += 1

    s1_headers = ["Employee Name", "Employee ID", "Hours", "Rate", "Base Pay",
                  f"Burden ({burden_pct}%)", "Total Cost"]
    for col, h in enumerate(s1_headers, 1):
        cell = ws.cell(row=s1_start, column=col, value=h)
        cell.font = s1_font
        cell.fill = s1_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Aggregate by employee (with OT-adjusted pay)
    emp_data = defaultdict(lambda: {"name": "", "emp_id": "", "hours": 0.0, "base": 0.0, "wage": None})
    for e in entries:
        key = e.get("emp_id_str", "")
        emp_data[key]["name"] = e.get("employee_name", "")
        emp_data[key]["emp_id"] = key
        emp_data[key]["hours"] += float(e.get("total_hours") or 0)
        emp_data[key]["base"] += _entry_base_pay(e)
        if e.get("hourly_wage") is not None:
            emp_data[key]["wage"] = e["hourly_wage"]

    sorted_emps = sorted(emp_data.values(), key=lambda x: x["name"].lower())
    total_hours = 0.0
    total_base = 0.0
    total_burden = 0.0
    total_cost = 0.0

    for i, emp in enumerate(sorted_emps):
        r = s1_start + 1 + i
        hours = round(emp["hours"], 2)
        total_hours += hours
        ws.cell(row=r, column=1, value=emp["name"]).border = thin_border
        ws.cell(row=r, column=2, value=emp["emp_id"]).border = thin_border
        c_hrs = ws.cell(row=r, column=3, value=hours)
        c_hrs.border = thin_border

        if emp["wage"] is not None:
            wage = emp["wage"]
            base = round(emp["base"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            total_base += base
            total_burden += burd
            total_cost += cost
            ws.cell(row=r, column=4, value=wage).border = thin_border
            ws.cell(row=r, column=4).number_format = money_fmt
            ws.cell(row=r, column=5, value=base).border = thin_border
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=6, value=burd).border = thin_border
            ws.cell(row=r, column=6).number_format = money_fmt
            ws.cell(row=r, column=7, value=cost).border = thin_border
            ws.cell(row=r, column=7).number_format = money_fmt
        else:
            for c in range(4, 8):
                ws.cell(row=r, column=c, value="—").border = thin_border

    tr = s1_start + 1 + len(sorted_emps)
    ws.cell(row=tr, column=1, value="Company Total").font = bold_font
    ws.cell(row=tr, column=1).border = thin_border
    ws.cell(row=tr, column=2).border = thin_border
    ws.cell(row=tr, column=3, value=round(total_hours, 2)).font = bold_font
    ws.cell(row=tr, column=3).border = thin_border
    ws.cell(row=tr, column=4).border = thin_border
    ws.cell(row=tr, column=5, value=round(total_base, 2)).font = bold_font
    ws.cell(row=tr, column=5).border = thin_border
    ws.cell(row=tr, column=5).number_format = money_fmt
    ws.cell(row=tr, column=6, value=round(total_burden, 2)).font = bold_font
    ws.cell(row=tr, column=6).border = thin_border
    ws.cell(row=tr, column=6).number_format = money_fmt
    ws.cell(row=tr, column=7, value=round(total_cost, 2)).font = bold_font
    ws.cell(row=tr, column=7).border = thin_border
    ws.cell(row=tr, column=7).number_format = money_fmt

    # --- Section 2: Employee Cost by Job ---
    s2_start = tr + 3
    s2_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    ws.cell(row=s2_start - 1, column=1, value="Employee Cost by Job").font = Font(bold=True, size=13)
    s2_headers = ["Employee Name", "Employee ID", "Job", "Hours", "Rate",
                  "Base Pay", "Burden", "Total Cost"]
    for col, h in enumerate(s2_headers, 1):
        cell = ws.cell(row=s2_start, column=col, value=h)
        cell.font = s1_font
        cell.fill = s2_fill
        cell.alignment = header_align
        cell.border = thin_border

    ej_data = defaultdict(lambda: {"name": "", "emp_id": "", "job": "", "hours": 0.0, "base": 0.0, "wage": None})
    for e in entries:
        key = (e.get("emp_id_str", ""), e.get("job_name", ""))
        ej_data[key]["name"] = e.get("employee_name", "")
        ej_data[key]["emp_id"] = e.get("emp_id_str", "")
        ej_data[key]["job"] = e.get("job_name", "")
        ej_data[key]["hours"] += float(e.get("total_hours") or 0)
        ej_data[key]["base"] += _entry_base_pay(e)
        if e.get("hourly_wage") is not None:
            ej_data[key]["wage"] = e["hourly_wage"]

    sorted_ej = sorted(ej_data.values(), key=lambda x: (x["name"].lower(), x["job"].lower()))
    for i, ej in enumerate(sorted_ej):
        r = s2_start + 1 + i
        hours = round(ej["hours"], 2)
        ws.cell(row=r, column=1, value=ej["name"]).border = thin_border
        ws.cell(row=r, column=2, value=ej["emp_id"]).border = thin_border
        ws.cell(row=r, column=3, value=ej["job"]).border = thin_border
        ws.cell(row=r, column=4, value=hours).border = thin_border
        if ej["wage"] is not None:
            wage = ej["wage"]
            base = round(ej["base"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            ws.cell(row=r, column=5, value=wage).border = thin_border
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=6, value=base).border = thin_border
            ws.cell(row=r, column=6).number_format = money_fmt
            ws.cell(row=r, column=7, value=burd).border = thin_border
            ws.cell(row=r, column=7).number_format = money_fmt
            ws.cell(row=r, column=8, value=cost).border = thin_border
            ws.cell(row=r, column=8).number_format = money_fmt
        else:
            for c in range(5, 9):
                ws.cell(row=r, column=c, value="—").border = thin_border

    # --- Section 3: Company Cost by Job ---
    s3_start = s2_start + 1 + len(sorted_ej) + 2
    s3_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    ws.cell(row=s3_start - 1, column=1, value="Company Cost by Job").font = Font(bold=True, size=13)
    s3_headers = ["Job", "Hours", "Base Pay", "Burden", "Total Cost"]
    for col, h in enumerate(s3_headers, 1):
        cell = ws.cell(row=s3_start, column=col, value=h)
        cell.font = s1_font
        cell.fill = s3_fill
        cell.alignment = header_align
        cell.border = thin_border

    job_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        jn = e.get("job_name", "")
        hours = float(e.get("total_hours") or 0)
        job_data[jn]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            job_data[jn]["base"] += base
            job_data[jn]["burden"] += burd
            job_data[jn]["cost"] += base + burd

    sorted_jobs = sorted(job_data.items(), key=lambda x: x[0].lower())
    jt_hours = jt_base = jt_burden = jt_cost = 0.0
    for i, (jname, jd) in enumerate(sorted_jobs):
        r = s3_start + 1 + i
        hrs = round(jd["hours"], 2)
        base = round(jd["base"], 2)
        burd = round(jd["burden"], 2)
        cost = round(jd["cost"], 2)
        jt_hours += hrs
        jt_base += base
        jt_burden += burd
        jt_cost += cost
        ws.cell(row=r, column=1, value=jname).border = thin_border
        ws.cell(row=r, column=2, value=hrs).border = thin_border
        ws.cell(row=r, column=3, value=base).border = thin_border
        ws.cell(row=r, column=3).number_format = money_fmt
        ws.cell(row=r, column=4, value=burd).border = thin_border
        ws.cell(row=r, column=4).number_format = money_fmt
        ws.cell(row=r, column=5, value=cost).border = thin_border
        ws.cell(row=r, column=5).number_format = money_fmt

    jtr = s3_start + 1 + len(sorted_jobs)
    ws.cell(row=jtr, column=1, value="Company Total").font = bold_font
    ws.cell(row=jtr, column=1).border = thin_border
    ws.cell(row=jtr, column=2, value=round(jt_hours, 2)).font = bold_font
    ws.cell(row=jtr, column=2).border = thin_border
    ws.cell(row=jtr, column=3, value=round(jt_base, 2)).font = bold_font
    ws.cell(row=jtr, column=3).border = thin_border
    ws.cell(row=jtr, column=3).number_format = money_fmt
    ws.cell(row=jtr, column=4, value=round(jt_burden, 2)).font = bold_font
    ws.cell(row=jtr, column=4).border = thin_border
    ws.cell(row=jtr, column=4).number_format = money_fmt
    ws.cell(row=jtr, column=5, value=round(jt_cost, 2)).font = bold_font
    ws.cell(row=jtr, column=5).border = thin_border
    ws.cell(row=jtr, column=5).number_format = money_fmt

    # --- Section 4: Company Cost by Date ---
    s4_start = jtr + 3
    s4_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ws.cell(row=s4_start - 1, column=1, value="Company Cost by Date").font = Font(bold=True, size=13)
    s4_headers = ["Date", "Hours", "Base Pay", "Burden", "Total Cost"]
    for col, h in enumerate(s4_headers, 1):
        cell = ws.cell(row=s4_start, column=col, value=h)
        cell.font = s1_font
        cell.fill = s4_fill
        cell.alignment = header_align
        cell.border = thin_border

    date_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        dt = (e.get("clock_in_time") or "")[:10]
        hours = float(e.get("total_hours") or 0)
        date_data[dt]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            date_data[dt]["base"] += base
            date_data[dt]["burden"] += burd
            date_data[dt]["cost"] += base + burd

    sorted_dates = sorted(date_data.items())
    dt_hours = dt_base = dt_burden = dt_cost = 0.0
    for i, (dt, dd) in enumerate(sorted_dates):
        r = s4_start + 1 + i
        hrs = round(dd["hours"], 2)
        base = round(dd["base"], 2)
        burd = round(dd["burden"], 2)
        cost = round(dd["cost"], 2)
        dt_hours += hrs
        dt_base += base
        dt_burden += burd
        dt_cost += cost
        ws.cell(row=r, column=1, value=dt).border = thin_border
        ws.cell(row=r, column=2, value=hrs).border = thin_border
        ws.cell(row=r, column=3, value=base).border = thin_border
        ws.cell(row=r, column=3).number_format = money_fmt
        ws.cell(row=r, column=4, value=burd).border = thin_border
        ws.cell(row=r, column=4).number_format = money_fmt
        ws.cell(row=r, column=5, value=cost).border = thin_border
        ws.cell(row=r, column=5).number_format = money_fmt

    dtr = s4_start + 1 + len(sorted_dates)
    ws.cell(row=dtr, column=1, value="Company Total").font = bold_font
    ws.cell(row=dtr, column=1).border = thin_border
    ws.cell(row=dtr, column=2, value=round(dt_hours, 2)).font = bold_font
    ws.cell(row=dtr, column=2).border = thin_border
    ws.cell(row=dtr, column=3, value=round(dt_base, 2)).font = bold_font
    ws.cell(row=dtr, column=3).border = thin_border
    ws.cell(row=dtr, column=3).number_format = money_fmt
    ws.cell(row=dtr, column=4, value=round(dt_burden, 2)).font = bold_font
    ws.cell(row=dtr, column=4).border = thin_border
    ws.cell(row=dtr, column=4).number_format = money_fmt
    ws.cell(row=dtr, column=5, value=round(dt_cost, 2)).font = bold_font
    ws.cell(row=dtr, column=5).border = thin_border
    ws.cell(row=dtr, column=5).number_format = money_fmt

    # Add logos
    _xl_add_logos(ws, token_str, dtr)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    for sheet in wb.worksheets:
        sheet.sheet_view.zoomScale = 140

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_range = ""
    if date_from:
        date_range += f"_{date_from}"
    if date_to:
        date_range += f"_to_{date_to}"
    filename = f"payroll_cost_{company.replace(' ', '_')}{date_range}.xlsx"

    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Combined Export (Hours + Payroll Cost)
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export/combined")
@login_required
def admin_export_combined():
    _app = _helpers()
    token_str = request.args.get("token", "")
    if not current_user.is_bdb:
        token_str = current_user.token
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    _app._verify_token_access(token_str)

    if not token_str:
        flash("Token is required.", "error")
        return redirect(url_for("time_admin.admin_export"))

    entries = database.get_time_entries_for_export(
        token_str, date_from=date_from or None, date_to=date_to or None,
    )
    token_data = database.get_token(token_str)
    company = token_data["company_name"] if token_data else "Unknown"
    burden_pct = token_data.get("labor_burden_pct", 0) if token_data else 0

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Report"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True, size=11)
    money_fmt = '#,##0.00'

    # Disclaimer rows
    ws.cell(row=1, column=1, value=f"COMBINED HOURS & PAYROLL REPORT — {company}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="NOT FOR BOOKKEEPING PURPOSES — Estimate Only").font = Font(bold=True, size=11, color="DC2626")
    range_label = ""
    if date_from:
        range_label += date_from
    if date_to:
        range_label += f" to {date_to}"
    ws.cell(row=3, column=1, value=f"Date range: {range_label}  |  Labor burden: {burden_pct}%").font = Font(size=10, color="6B7280")

    # Pre-compute OT effective rates for all entries
    eff_rates = database.get_effective_rates_for_entries(token_str, entries)

    def _entry_base_pay(e):
        """Return OT-adjusted base pay for a single time entry."""
        hrs = float(e.get("total_hours") or 0)
        if hrs <= 0 or e.get("hourly_wage") is None:
            return 0.0
        week = database._get_week_start_sunday(e["clock_in_time"])
        rate_info = eff_rates.get((e["employee_id"], week))
        if rate_info and rate_info["effective_rate"]:
            return hrs * rate_info["effective_rate"]
        return hrs * e["hourly_wage"]

    # --- Section 1: Employee Summary (green) ---
    s1_start = 5
    s1_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    ws.cell(row=s1_start, column=1, value="Employee Summary").font = Font(bold=True, size=13)
    s1_start += 1

    s1_headers = ["Employee Name", "Employee ID", "Total Hours", "Rate",
                  "Base Pay", f"Burden ({burden_pct}%)", "Total Cost"]
    for col, h in enumerate(s1_headers, 1):
        cell = ws.cell(row=s1_start, column=col, value=h)
        cell.font = header_font
        cell.fill = s1_fill
        cell.alignment = header_align
        cell.border = thin_border

    emp_data = defaultdict(lambda: {"name": "", "emp_id": "", "hours": 0.0, "base": 0.0, "wage": None})
    for e in entries:
        key = e.get("emp_id_str", "")
        emp_data[key]["name"] = e.get("employee_name", "")
        emp_data[key]["emp_id"] = key
        emp_data[key]["hours"] += float(e.get("total_hours") or 0)
        emp_data[key]["base"] += _entry_base_pay(e)
        if e.get("hourly_wage") is not None:
            emp_data[key]["wage"] = e["hourly_wage"]

    sorted_emps = sorted(emp_data.values(), key=lambda x: x["name"].lower())
    company_hours = company_base = company_burden = company_cost = 0.0

    for i, emp in enumerate(sorted_emps):
        r = s1_start + 1 + i
        hours = round(emp["hours"], 2)
        company_hours += hours
        ws.cell(row=r, column=1, value=emp["name"]).border = thin_border
        ws.cell(row=r, column=2, value=emp["emp_id"]).border = thin_border
        ws.cell(row=r, column=3, value=hours).border = thin_border
        if emp["wage"] is not None:
            wage = emp["wage"]
            base = round(emp["base"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            company_base += base
            company_burden += burd
            company_cost += cost
            ws.cell(row=r, column=4, value=wage).border = thin_border
            ws.cell(row=r, column=4).number_format = money_fmt
            ws.cell(row=r, column=5, value=base).border = thin_border
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=6, value=burd).border = thin_border
            ws.cell(row=r, column=6).number_format = money_fmt
            ws.cell(row=r, column=7, value=cost).border = thin_border
            ws.cell(row=r, column=7).number_format = money_fmt
        else:
            for c in range(4, 8):
                ws.cell(row=r, column=c, value="—").border = thin_border

    tr = s1_start + 1 + len(sorted_emps)
    ws.cell(row=tr, column=1, value="Company Total").font = bold_font
    ws.cell(row=tr, column=1).border = thin_border
    ws.cell(row=tr, column=2).border = thin_border
    ws.cell(row=tr, column=3, value=round(company_hours, 2)).font = bold_font
    ws.cell(row=tr, column=3).border = thin_border
    ws.cell(row=tr, column=4).border = thin_border
    ws.cell(row=tr, column=5, value=round(company_base, 2)).font = bold_font
    ws.cell(row=tr, column=5).border = thin_border
    ws.cell(row=tr, column=5).number_format = money_fmt
    ws.cell(row=tr, column=6, value=round(company_burden, 2)).font = bold_font
    ws.cell(row=tr, column=6).border = thin_border
    ws.cell(row=tr, column=6).number_format = money_fmt
    ws.cell(row=tr, column=7, value=round(company_cost, 2)).font = bold_font
    ws.cell(row=tr, column=7).border = thin_border
    ws.cell(row=tr, column=7).number_format = money_fmt

    # --- Section 2: Employee Hours by Job + Cost (orange) ---
    s2_start = tr + 3
    s2_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    ws.cell(row=s2_start - 1, column=1, value="Employee Hours by Job").font = Font(bold=True, size=13)
    s2_headers = ["Employee Name", "Employee ID", "Job", "Hours", "Rate",
                  "Base Pay", "Burden", "Total Cost"]
    for col, h in enumerate(s2_headers, 1):
        cell = ws.cell(row=s2_start, column=col, value=h)
        cell.font = header_font
        cell.fill = s2_fill
        cell.alignment = header_align
        cell.border = thin_border

    ej_data = defaultdict(lambda: {"name": "", "emp_id": "", "job": "", "hours": 0.0, "base": 0.0, "wage": None})
    for e in entries:
        key = (e.get("emp_id_str", ""), e.get("job_name", ""))
        ej_data[key]["name"] = e.get("employee_name", "")
        ej_data[key]["emp_id"] = e.get("emp_id_str", "")
        ej_data[key]["job"] = e.get("job_name", "")
        ej_data[key]["hours"] += float(e.get("total_hours") or 0)
        ej_data[key]["base"] += _entry_base_pay(e)
        if e.get("hourly_wage") is not None:
            ej_data[key]["wage"] = e["hourly_wage"]

    sorted_ej = sorted(ej_data.values(), key=lambda x: (x["name"].lower(), x["job"].lower()))
    for i, ej in enumerate(sorted_ej):
        r = s2_start + 1 + i
        hours = round(ej["hours"], 2)
        ws.cell(row=r, column=1, value=ej["name"]).border = thin_border
        ws.cell(row=r, column=2, value=ej["emp_id"]).border = thin_border
        ws.cell(row=r, column=3, value=ej["job"]).border = thin_border
        ws.cell(row=r, column=4, value=hours).border = thin_border
        if ej["wage"] is not None:
            base = round(ej["base"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            ws.cell(row=r, column=5, value=ej["wage"]).border = thin_border
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=6, value=base).border = thin_border
            ws.cell(row=r, column=6).number_format = money_fmt
            ws.cell(row=r, column=7, value=burd).border = thin_border
            ws.cell(row=r, column=7).number_format = money_fmt
            ws.cell(row=r, column=8, value=cost).border = thin_border
            ws.cell(row=r, column=8).number_format = money_fmt
        else:
            for c in range(5, 9):
                ws.cell(row=r, column=c, value="—").border = thin_border

    # --- Section 3: Company Hours by Job + Cost (purple) ---
    s3_start = s2_start + 1 + len(sorted_ej) + 2
    s3_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    ws.cell(row=s3_start - 1, column=1, value="Company Hours by Job").font = Font(bold=True, size=13)
    s3_headers = ["Job", "Hours", "Base Pay", "Burden", "Total Cost"]
    for col, h in enumerate(s3_headers, 1):
        cell = ws.cell(row=s3_start, column=col, value=h)
        cell.font = header_font
        cell.fill = s3_fill
        cell.alignment = header_align
        cell.border = thin_border

    job_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        jn = e.get("job_name", "")
        hours = float(e.get("total_hours") or 0)
        job_data[jn]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            job_data[jn]["base"] += base
            job_data[jn]["burden"] += burd
            job_data[jn]["cost"] += base + burd

    sorted_jobs = sorted(job_data.items(), key=lambda x: x[0].lower())
    jt_hours = jt_base = jt_burden = jt_cost = 0.0
    for i, (jname, jd) in enumerate(sorted_jobs):
        r = s3_start + 1 + i
        hrs = round(jd["hours"], 2)
        base = round(jd["base"], 2)
        burd = round(jd["burden"], 2)
        cost = round(jd["cost"], 2)
        jt_hours += hrs
        jt_base += base
        jt_burden += burd
        jt_cost += cost
        ws.cell(row=r, column=1, value=jname).border = thin_border
        ws.cell(row=r, column=2, value=hrs).border = thin_border
        ws.cell(row=r, column=3, value=base).border = thin_border
        ws.cell(row=r, column=3).number_format = money_fmt
        ws.cell(row=r, column=4, value=burd).border = thin_border
        ws.cell(row=r, column=4).number_format = money_fmt
        ws.cell(row=r, column=5, value=cost).border = thin_border
        ws.cell(row=r, column=5).number_format = money_fmt

    jtr = s3_start + 1 + len(sorted_jobs)
    ws.cell(row=jtr, column=1, value="Company Total").font = bold_font
    ws.cell(row=jtr, column=1).border = thin_border
    ws.cell(row=jtr, column=2, value=round(jt_hours, 2)).font = bold_font
    ws.cell(row=jtr, column=2).border = thin_border
    ws.cell(row=jtr, column=3, value=round(jt_base, 2)).font = bold_font
    ws.cell(row=jtr, column=3).border = thin_border
    ws.cell(row=jtr, column=3).number_format = money_fmt
    ws.cell(row=jtr, column=4, value=round(jt_burden, 2)).font = bold_font
    ws.cell(row=jtr, column=4).border = thin_border
    ws.cell(row=jtr, column=4).number_format = money_fmt
    ws.cell(row=jtr, column=5, value=round(jt_cost, 2)).font = bold_font
    ws.cell(row=jtr, column=5).border = thin_border
    ws.cell(row=jtr, column=5).number_format = money_fmt

    # --- Section 4: Company Cost by Date (blue) ---
    s4_start = jtr + 3
    s4_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ws.cell(row=s4_start - 1, column=1, value="Company Cost by Date").font = Font(bold=True, size=13)
    s4_headers = ["Date", "Hours", "Base Pay", "Burden", "Total Cost"]
    for col, h in enumerate(s4_headers, 1):
        cell = ws.cell(row=s4_start, column=col, value=h)
        cell.font = header_font
        cell.fill = s4_fill
        cell.alignment = header_align
        cell.border = thin_border

    date_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        dt = (e.get("clock_in_time") or "")[:10]
        hours = float(e.get("total_hours") or 0)
        date_data[dt]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            date_data[dt]["base"] += base
            date_data[dt]["burden"] += burd
            date_data[dt]["cost"] += base + burd

    sorted_dates = sorted(date_data.items())
    dt_hours = dt_base = dt_burden = dt_cost = 0.0
    for i, (dt, dd) in enumerate(sorted_dates):
        r = s4_start + 1 + i
        hrs = round(dd["hours"], 2)
        base = round(dd["base"], 2)
        burd = round(dd["burden"], 2)
        cost = round(dd["cost"], 2)
        dt_hours += hrs
        dt_base += base
        dt_burden += burd
        dt_cost += cost
        ws.cell(row=r, column=1, value=dt).border = thin_border
        ws.cell(row=r, column=2, value=hrs).border = thin_border
        ws.cell(row=r, column=3, value=base).border = thin_border
        ws.cell(row=r, column=3).number_format = money_fmt
        ws.cell(row=r, column=4, value=burd).border = thin_border
        ws.cell(row=r, column=4).number_format = money_fmt
        ws.cell(row=r, column=5, value=cost).border = thin_border
        ws.cell(row=r, column=5).number_format = money_fmt

    dtr = s4_start + 1 + len(sorted_dates)
    ws.cell(row=dtr, column=1, value="Company Total").font = bold_font
    ws.cell(row=dtr, column=1).border = thin_border
    ws.cell(row=dtr, column=2, value=round(dt_hours, 2)).font = bold_font
    ws.cell(row=dtr, column=2).border = thin_border
    ws.cell(row=dtr, column=3, value=round(dt_base, 2)).font = bold_font
    ws.cell(row=dtr, column=3).border = thin_border
    ws.cell(row=dtr, column=3).number_format = money_fmt
    ws.cell(row=dtr, column=4, value=round(dt_burden, 2)).font = bold_font
    ws.cell(row=dtr, column=4).border = thin_border
    ws.cell(row=dtr, column=4).number_format = money_fmt
    ws.cell(row=dtr, column=5, value=round(dt_cost, 2)).font = bold_font
    ws.cell(row=dtr, column=5).border = thin_border
    ws.cell(row=dtr, column=5).number_format = money_fmt

    # Add logos
    _xl_add_logos(ws, token_str, dtr)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    for sheet in wb.worksheets:
        sheet.sheet_view.zoomScale = 140

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_range = ""
    if date_from:
        date_range += f"_{date_from}"
    if date_to:
        date_range += f"_to_{date_to}"
    filename = f"combined_{company.replace(' ', '_')}{date_range}.xlsx"

    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Payroll Cost PDF Export
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export/payroll-cost-pdf")
@login_required
def admin_export_payroll_cost_pdf():
    _app = _helpers()
    token_str = request.args.get("token", "")
    if not current_user.is_bdb:
        token_str = current_user.token
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    _app._verify_token_access(token_str)

    if not token_str:
        flash("Token is required.", "error")
        return redirect(url_for("time_admin.admin_export"))

    entries = database.get_time_entries_for_export(
        token_str, date_from=date_from or None, date_to=date_to or None,
    )
    token_data = database.get_token(token_str)
    company = token_data["company_name"] if token_data else "Unknown"
    burden_pct = token_data.get("labor_burden_pct", 0) if token_data else 0
    company_logo = _company_logo_path(token_str)

    eff_rates = database.get_effective_rates_for_entries(token_str, entries)

    def _entry_base_pay(e):
        hrs = float(e.get("total_hours") or 0)
        if hrs <= 0 or e.get("hourly_wage") is None:
            return 0.0
        week = database._get_week_start_sunday(e["clock_in_time"])
        rate_info = eff_rates.get((e["employee_id"], week))
        if rate_info and rate_info["effective_rate"]:
            return hrs * rate_info["effective_rate"]
        return hrs * e["hourly_wage"]

    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=28)
    pdf.add_page()

    _pdf_add_logos(pdf, company_logo)

    # Title
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _safe(f"{company} - Payroll Cost Report"), ln=True, align="C")
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(220, 38, 38)
    pdf.cell(0, 6, "NOT FOR BOOKKEEPING PURPOSES", ln=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)
    range_label = ""
    if date_from:
        range_label += date_from
    if date_to:
        range_label += f"  to  {date_to}"
    if range_label:
        pdf.cell(0, 5, _safe(f"{range_label}  |  Labor burden: {burden_pct}%"), ln=True, align="C")
    pdf.ln(4)

    # Aggregate employee data
    emp_data = defaultdict(lambda: {"name": "", "emp_id": "", "hours": 0.0, "wage": None})
    for e in entries:
        key = e.get("emp_id_str", "")
        emp_data[key]["name"] = e.get("employee_name", "")
        emp_data[key]["emp_id"] = key
        emp_data[key]["hours"] += float(e.get("total_hours") or 0)
        if e.get("hourly_wage") is not None:
            emp_data[key]["wage"] = e["hourly_wage"]
    sorted_emps = sorted(emp_data.values(), key=lambda x: x["name"].lower())

    # --- Section 1: Employee Cost Summary (green) ---
    _pdf_section_header(pdf, "Employee Cost Summary", _SECTION_COLORS["green"])
    s1_widths = [50, 30, 25, 25, 30, 30, 30]
    _pdf_table_header(pdf, ["Employee", "Emp ID", "Hours", "Rate", "Base Pay",
                            f"Burden ({burden_pct}%)", "Total Cost"],
                      s1_widths, _SECTION_COLORS["green"])

    total_hours = total_base = total_burden = total_cost = 0.0
    pdf.set_font("Helvetica", "", 8)
    for emp in sorted_emps:
        hours = round(emp["hours"], 2)
        total_hours += hours
        pdf.cell(s1_widths[0], 6, _safe(emp["name"][:30]), border=1)
        pdf.cell(s1_widths[1], 6, _safe(emp["emp_id"]), border=1)
        pdf.cell(s1_widths[2], 6, str(hours), border=1, align="R")
        if emp["wage"] is not None:
            base = round(hours * emp["wage"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            total_base += base
            total_burden += burd
            total_cost += cost
            pdf.cell(s1_widths[3], 6, f"${emp['wage']:.2f}", border=1, align="R")
            pdf.cell(s1_widths[4], 6, f"${base:,.2f}", border=1, align="R")
            pdf.cell(s1_widths[5], 6, f"${burd:,.2f}", border=1, align="R")
            pdf.cell(s1_widths[6], 6, f"${cost:,.2f}", border=1, align="R")
        else:
            for _ in range(4):
                pdf.cell(s1_widths[3], 6, _safe("--"), border=1, align="C")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(s1_widths[0] + s1_widths[1], 7, "Company Total", border=1)
    pdf.cell(s1_widths[2], 7, str(round(total_hours, 2)), border=1, align="R")
    pdf.cell(s1_widths[3], 7, "", border=1)
    pdf.cell(s1_widths[4], 7, f"${total_base:,.2f}", border=1, align="R")
    pdf.cell(s1_widths[5], 7, f"${total_burden:,.2f}", border=1, align="R")
    pdf.cell(s1_widths[6], 7, f"${total_cost:,.2f}", border=1, align="R")
    pdf.ln()

    # --- Section 2: Employee Cost by Job (orange) ---
    _pdf_section_header(pdf, "Employee Cost by Job", _SECTION_COLORS["orange"])
    s2_widths = [42, 25, 45, 22, 22, 28, 28, 28]
    _pdf_table_header(pdf, ["Employee", "Emp ID", "Job", "Hours", "Rate",
                            "Base Pay", "Burden", "Total Cost"],
                      s2_widths, _SECTION_COLORS["orange"])

    ej_data = defaultdict(lambda: {"name": "", "emp_id": "", "job": "", "hours": 0.0, "wage": None})
    for e in entries:
        key = (e.get("emp_id_str", ""), e.get("job_name", ""))
        ej_data[key]["name"] = e.get("employee_name", "")
        ej_data[key]["emp_id"] = e.get("emp_id_str", "")
        ej_data[key]["job"] = e.get("job_name", "")
        ej_data[key]["hours"] += float(e.get("total_hours") or 0)
        if e.get("hourly_wage") is not None:
            ej_data[key]["wage"] = e["hourly_wage"]

    sorted_ej = sorted(ej_data.values(), key=lambda x: (x["name"].lower(), x["job"].lower()))
    pdf.set_font("Helvetica", "", 8)
    for ej in sorted_ej:
        hours = round(ej["hours"], 2)
        pdf.cell(s2_widths[0], 6, _safe(ej["name"][:25]), border=1)
        pdf.cell(s2_widths[1], 6, _safe(ej["emp_id"]), border=1)
        pdf.cell(s2_widths[2], 6, _safe(ej["job"][:28]), border=1)
        pdf.cell(s2_widths[3], 6, str(hours), border=1, align="R")
        if ej["wage"] is not None:
            base = round(hours * ej["wage"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            pdf.cell(s2_widths[4], 6, f"${ej['wage']:.2f}", border=1, align="R")
            pdf.cell(s2_widths[5], 6, f"${base:,.2f}", border=1, align="R")
            pdf.cell(s2_widths[6], 6, f"${burd:,.2f}", border=1, align="R")
            pdf.cell(s2_widths[7], 6, f"${cost:,.2f}", border=1, align="R")
        else:
            for _ in range(4):
                pdf.cell(s2_widths[4], 6, _safe("--"), border=1, align="C")
        pdf.ln()

    # --- Section 3: Company Cost by Job (purple) ---
    _pdf_section_header(pdf, "Company Cost by Job", _SECTION_COLORS["purple"])
    s3_widths = [80, 25, 35, 35, 35]
    _pdf_table_header(pdf, ["Job", "Hours", "Base Pay", "Burden", "Total Cost"],
                      s3_widths, _SECTION_COLORS["purple"])

    job_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        jn = e.get("job_name", "")
        hours = float(e.get("total_hours") or 0)
        job_data[jn]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            job_data[jn]["base"] += base
            job_data[jn]["burden"] += burd
            job_data[jn]["cost"] += base + burd

    sorted_jobs = sorted(job_data.items(), key=lambda x: x[0].lower())
    pdf.set_font("Helvetica", "", 8)
    for jname, jd in sorted_jobs:
        pdf.cell(s3_widths[0], 6, _safe(jname[:48]), border=1)
        pdf.cell(s3_widths[1], 6, str(round(jd["hours"], 2)), border=1, align="R")
        pdf.cell(s3_widths[2], 6, f"${jd['base']:,.2f}", border=1, align="R")
        pdf.cell(s3_widths[3], 6, f"${jd['burden']:,.2f}", border=1, align="R")
        pdf.cell(s3_widths[4], 6, f"${jd['cost']:,.2f}", border=1, align="R")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(s3_widths[0], 7, "Company Total", border=1)
    pdf.cell(s3_widths[1], 7, str(round(total_hours, 2)), border=1, align="R")
    pdf.cell(s3_widths[2], 7, f"${total_base:,.2f}", border=1, align="R")
    pdf.cell(s3_widths[3], 7, f"${total_burden:,.2f}", border=1, align="R")
    pdf.cell(s3_widths[4], 7, f"${total_cost:,.2f}", border=1, align="R")
    pdf.ln()

    # --- Section 4: Company Cost by Date (blue) ---
    _pdf_section_header(pdf, "Company Cost by Date", _SECTION_COLORS["blue"])
    s4_widths = [80, 25, 35, 35, 35]
    _pdf_table_header(pdf, ["Date", "Hours", "Base Pay", "Burden", "Total Cost"],
                      s4_widths, _SECTION_COLORS["blue"])

    date_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        dt = (e.get("clock_in_time") or "")[:10]
        hours = float(e.get("total_hours") or 0)
        date_data[dt]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            date_data[dt]["base"] += base
            date_data[dt]["burden"] += burd
            date_data[dt]["cost"] += base + burd

    sorted_dates = sorted(date_data.items())
    pdf.set_font("Helvetica", "", 8)
    for dt, dd in sorted_dates:
        pdf.cell(s4_widths[0], 6, _safe(dt), border=1)
        pdf.cell(s4_widths[1], 6, str(round(dd["hours"], 2)), border=1, align="R")
        pdf.cell(s4_widths[2], 6, f"${dd['base']:,.2f}", border=1, align="R")
        pdf.cell(s4_widths[3], 6, f"${dd['burden']:,.2f}", border=1, align="R")
        pdf.cell(s4_widths[4], 6, f"${dd['cost']:,.2f}", border=1, align="R")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(s4_widths[0], 7, "Company Total", border=1)
    pdf.cell(s4_widths[1], 7, str(round(total_hours, 2)), border=1, align="R")
    pdf.cell(s4_widths[2], 7, f"${total_base:,.2f}", border=1, align="R")
    pdf.cell(s4_widths[3], 7, f"${total_burden:,.2f}", border=1, align="R")
    pdf.cell(s4_widths[4], 7, f"${total_cost:,.2f}", border=1, align="R")
    pdf.ln()

    _pdf_add_bdb_footer(pdf)

    output = BytesIO()
    pdf.output(output)
    output.seek(0)

    date_range = ""
    if date_from:
        date_range += f"_{date_from}"
    if date_to:
        date_range += f"_to_{date_to}"
    filename = f"payroll_cost_{company.replace(' ', '_')}{date_range}.pdf"

    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/pdf",
    )


# ---------------------------------------------------------------------------
# Combined PDF Export
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/export/combined-pdf")
@login_required
def admin_export_combined_pdf():
    _app = _helpers()
    token_str = request.args.get("token", "")
    if not current_user.is_bdb:
        token_str = current_user.token
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    _app._verify_token_access(token_str)

    if not token_str:
        flash("Token is required.", "error")
        return redirect(url_for("time_admin.admin_export"))

    entries = database.get_time_entries_for_export(
        token_str, date_from=date_from or None, date_to=date_to or None,
    )
    token_data = database.get_token(token_str)
    company = token_data["company_name"] if token_data else "Unknown"
    burden_pct = token_data.get("labor_burden_pct", 0) if token_data else 0
    company_logo = _company_logo_path(token_str)

    eff_rates = database.get_effective_rates_for_entries(token_str, entries)

    def _entry_base_pay(e):
        hrs = float(e.get("total_hours") or 0)
        if hrs <= 0 or e.get("hourly_wage") is None:
            return 0.0
        week = database._get_week_start_sunday(e["clock_in_time"])
        rate_info = eff_rates.get((e["employee_id"], week))
        if rate_info and rate_info["effective_rate"]:
            return hrs * rate_info["effective_rate"]
        return hrs * e["hourly_wage"]

    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=28)
    pdf.add_page()

    _pdf_add_logos(pdf, company_logo)

    # Title
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _safe(f"{company} - Combined Report"), ln=True, align="C")
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(220, 38, 38)
    pdf.cell(0, 6, "NOT FOR BOOKKEEPING PURPOSES", ln=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)
    range_label = ""
    if date_from:
        range_label += date_from
    if date_to:
        range_label += f"  to  {date_to}"
    if range_label:
        pdf.cell(0, 5, _safe(f"{range_label}  |  Labor burden: {burden_pct}%"), ln=True, align="C")
    pdf.ln(4)

    # Aggregate data
    emp_data = defaultdict(lambda: {"name": "", "emp_id": "", "hours": 0.0, "wage": None})
    for e in entries:
        key = e.get("emp_id_str", "")
        emp_data[key]["name"] = e.get("employee_name", "")
        emp_data[key]["emp_id"] = key
        emp_data[key]["hours"] += float(e.get("total_hours") or 0)
        if e.get("hourly_wage") is not None:
            emp_data[key]["wage"] = e["hourly_wage"]
    sorted_emps = sorted(emp_data.values(), key=lambda x: x["name"].lower())

    # --- Section 1: Employee Summary (green) ---
    _pdf_section_header(pdf, "Employee Summary", _SECTION_COLORS["green"])
    s1_widths = [50, 30, 25, 25, 30, 30, 30]
    _pdf_table_header(pdf, ["Employee", "Emp ID", "Hours", "Rate", "Base Pay",
                            f"Burden ({burden_pct}%)", "Total Cost"],
                      s1_widths, _SECTION_COLORS["green"])

    company_hours = company_base = company_burden = company_cost = 0.0
    pdf.set_font("Helvetica", "", 8)
    for emp in sorted_emps:
        hours = round(emp["hours"], 2)
        company_hours += hours
        pdf.cell(s1_widths[0], 6, _safe(emp["name"][:30]), border=1)
        pdf.cell(s1_widths[1], 6, _safe(emp["emp_id"]), border=1)
        pdf.cell(s1_widths[2], 6, str(hours), border=1, align="R")
        if emp["wage"] is not None:
            base = round(hours * emp["wage"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            company_base += base
            company_burden += burd
            company_cost += cost
            pdf.cell(s1_widths[3], 6, f"${emp['wage']:.2f}", border=1, align="R")
            pdf.cell(s1_widths[4], 6, f"${base:,.2f}", border=1, align="R")
            pdf.cell(s1_widths[5], 6, f"${burd:,.2f}", border=1, align="R")
            pdf.cell(s1_widths[6], 6, f"${cost:,.2f}", border=1, align="R")
        else:
            for _ in range(4):
                pdf.cell(s1_widths[3], 6, _safe("--"), border=1, align="C")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(s1_widths[0] + s1_widths[1], 7, "Company Total", border=1)
    pdf.cell(s1_widths[2], 7, str(round(company_hours, 2)), border=1, align="R")
    pdf.cell(s1_widths[3], 7, "", border=1)
    pdf.cell(s1_widths[4], 7, f"${company_base:,.2f}", border=1, align="R")
    pdf.cell(s1_widths[5], 7, f"${company_burden:,.2f}", border=1, align="R")
    pdf.cell(s1_widths[6], 7, f"${company_cost:,.2f}", border=1, align="R")
    pdf.ln()

    # --- Section 2: Employee Hours by Job (orange) ---
    _pdf_section_header(pdf, "Employee Hours by Job", _SECTION_COLORS["orange"])
    s2_widths = [42, 25, 45, 22, 22, 28, 28, 28]
    _pdf_table_header(pdf, ["Employee", "Emp ID", "Job", "Hours", "Rate",
                            "Base Pay", "Burden", "Total Cost"],
                      s2_widths, _SECTION_COLORS["orange"])

    ej_data = defaultdict(lambda: {"name": "", "emp_id": "", "job": "", "hours": 0.0, "wage": None})
    for e in entries:
        key = (e.get("emp_id_str", ""), e.get("job_name", ""))
        ej_data[key]["name"] = e.get("employee_name", "")
        ej_data[key]["emp_id"] = e.get("emp_id_str", "")
        ej_data[key]["job"] = e.get("job_name", "")
        ej_data[key]["hours"] += float(e.get("total_hours") or 0)
        if e.get("hourly_wage") is not None:
            ej_data[key]["wage"] = e["hourly_wage"]

    sorted_ej = sorted(ej_data.values(), key=lambda x: (x["name"].lower(), x["job"].lower()))
    pdf.set_font("Helvetica", "", 8)
    for ej in sorted_ej:
        hours = round(ej["hours"], 2)
        pdf.cell(s2_widths[0], 6, _safe(ej["name"][:25]), border=1)
        pdf.cell(s2_widths[1], 6, _safe(ej["emp_id"]), border=1)
        pdf.cell(s2_widths[2], 6, _safe(ej["job"][:28]), border=1)
        pdf.cell(s2_widths[3], 6, str(hours), border=1, align="R")
        if ej["wage"] is not None:
            base = round(hours * ej["wage"], 2)
            burd = round(base * (burden_pct / 100), 2)
            cost = round(base + burd, 2)
            pdf.cell(s2_widths[4], 6, f"${ej['wage']:.2f}", border=1, align="R")
            pdf.cell(s2_widths[5], 6, f"${base:,.2f}", border=1, align="R")
            pdf.cell(s2_widths[6], 6, f"${burd:,.2f}", border=1, align="R")
            pdf.cell(s2_widths[7], 6, f"${cost:,.2f}", border=1, align="R")
        else:
            for _ in range(4):
                pdf.cell(s2_widths[4], 6, _safe("--"), border=1, align="C")
        pdf.ln()

    # --- Section 3: Company Hours by Job (purple) ---
    _pdf_section_header(pdf, "Company Hours by Job", _SECTION_COLORS["purple"])
    s3_widths = [80, 25, 35, 35, 35]
    _pdf_table_header(pdf, ["Job", "Hours", "Base Pay", "Burden", "Total Cost"],
                      s3_widths, _SECTION_COLORS["purple"])

    job_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        jn = e.get("job_name", "")
        hours = float(e.get("total_hours") or 0)
        job_data[jn]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            job_data[jn]["base"] += base
            job_data[jn]["burden"] += burd
            job_data[jn]["cost"] += base + burd

    sorted_jobs = sorted(job_data.items(), key=lambda x: x[0].lower())
    jt_hours = jt_base = jt_burden = jt_cost = 0.0
    pdf.set_font("Helvetica", "", 8)
    for jname, jd in sorted_jobs:
        jt_hours += jd["hours"]
        jt_base += jd["base"]
        jt_burden += jd["burden"]
        jt_cost += jd["cost"]
        pdf.cell(s3_widths[0], 6, _safe(jname[:48]), border=1)
        pdf.cell(s3_widths[1], 6, str(round(jd["hours"], 2)), border=1, align="R")
        pdf.cell(s3_widths[2], 6, f"${jd['base']:,.2f}", border=1, align="R")
        pdf.cell(s3_widths[3], 6, f"${jd['burden']:,.2f}", border=1, align="R")
        pdf.cell(s3_widths[4], 6, f"${jd['cost']:,.2f}", border=1, align="R")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(s3_widths[0], 7, "Company Total", border=1)
    pdf.cell(s3_widths[1], 7, str(round(jt_hours, 2)), border=1, align="R")
    pdf.cell(s3_widths[2], 7, f"${jt_base:,.2f}", border=1, align="R")
    pdf.cell(s3_widths[3], 7, f"${jt_burden:,.2f}", border=1, align="R")
    pdf.cell(s3_widths[4], 7, f"${jt_cost:,.2f}", border=1, align="R")
    pdf.ln()

    # --- Section 4: Company Cost by Date (blue) ---
    _pdf_section_header(pdf, "Company Cost by Date", _SECTION_COLORS["blue"])
    s4_widths = [80, 25, 35, 35, 35]
    _pdf_table_header(pdf, ["Date", "Hours", "Base Pay", "Burden", "Total Cost"],
                      s4_widths, _SECTION_COLORS["blue"])

    date_data = defaultdict(lambda: {"hours": 0.0, "base": 0.0, "burden": 0.0, "cost": 0.0})
    for e in entries:
        dt = (e.get("clock_in_time") or "")[:10]
        hours = float(e.get("total_hours") or 0)
        date_data[dt]["hours"] += hours
        base = _entry_base_pay(e)
        if base > 0:
            burd = base * (burden_pct / 100)
            date_data[dt]["base"] += base
            date_data[dt]["burden"] += burd
            date_data[dt]["cost"] += base + burd

    sorted_dates = sorted(date_data.items())
    pdf.set_font("Helvetica", "", 8)
    for dt, dd in sorted_dates:
        pdf.cell(s4_widths[0], 6, _safe(dt), border=1)
        pdf.cell(s4_widths[1], 6, str(round(dd["hours"], 2)), border=1, align="R")
        pdf.cell(s4_widths[2], 6, f"${dd['base']:,.2f}", border=1, align="R")
        pdf.cell(s4_widths[3], 6, f"${dd['burden']:,.2f}", border=1, align="R")
        pdf.cell(s4_widths[4], 6, f"${dd['cost']:,.2f}", border=1, align="R")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(s4_widths[0], 7, "Company Total", border=1)
    pdf.cell(s4_widths[1], 7, str(round(company_hours, 2)), border=1, align="R")
    pdf.cell(s4_widths[2], 7, f"${company_base:,.2f}", border=1, align="R")
    pdf.cell(s4_widths[3], 7, f"${company_burden:,.2f}", border=1, align="R")
    pdf.cell(s4_widths[4], 7, f"${company_cost:,.2f}", border=1, align="R")
    pdf.ln()

    _pdf_add_bdb_footer(pdf)

    output = BytesIO()
    pdf.output(output)
    output.seek(0)

    date_range = ""
    if date_from:
        date_range += f"_{date_from}"
    if date_to:
        date_range += f"_to_{date_to}"
    filename = f"combined_{company.replace(' ', '_')}{date_range}.pdf"

    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/pdf",
    )


# ---------------------------------------------------------------------------
# Admin Guide
# ---------------------------------------------------------------------------

@time_admin_bp.route("/admin/guide")
@login_required
def admin_guide():
    _app = _helpers()
    tokens = _app._get_tokens_for_user()
    token_str, selected_token = _app._get_selected_token(tokens)
    return render_template("admin/guide.html", tokens=tokens, selected_token=selected_token)
