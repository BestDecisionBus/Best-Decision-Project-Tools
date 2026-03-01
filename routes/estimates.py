import io
import re
import uuid
from datetime import datetime
from pathlib import Path

from flask import (
    Blueprint, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for,
)
from flask_login import current_user

import config
import database

estimates_bp = Blueprint('estimates', __name__)


def _compute_finance_targets(token_str, token_data):
    """Return (margin_target, markup_required) matching the finance dashboard logic."""
    income_target_pct = (token_data or {}).get("income_target_pct", 0) or 0
    monthly_overhead = (token_data or {}).get("monthly_overhead", 0) or 0
    try:
        job_financials = database.get_job_financials(token_str)
        t_earned = sum(
            jf["budget"] * jf["completion_pct"] / 100
            for jf in job_financials
        )
        overhead_pct = round((monthly_overhead * 12) / t_earned * 100, 1) if (monthly_overhead > 0 and t_earned > 0) else 0
    except Exception:
        overhead_pct = 0
    margin_target = round(overhead_pct + income_target_pct, 1)
    markup_required = round(margin_target / (100 - margin_target) * 100, 1) if margin_target < 100 else 0
    return margin_target, markup_required


# ---------------------------------------------------------------------------
# Lazy import of app-level helpers to avoid circular imports
# ---------------------------------------------------------------------------

def _helpers():
    import app as _app
    return _app


def _require_admin(h=None):
    """Abort 403 if the current user is not an admin or BDB user."""
    user = h.current_user if h else current_user
    if not user.is_admin and not user.is_bdb:
        abort(403)


# ---------------------------------------------------------------------------
# Audio file validation
# ---------------------------------------------------------------------------

AUDIO_SIGNATURES = [
    b"\x1a\x45\xdf\xa3",  # WebM
    b"OggS",               # OGG
    b"\xff\xfb",           # MP3
    b"\xff\xf3",           # MP3
    b"\xff\xf2",           # MP3
    b"ID3",                # MP3 with ID3 tag
    b"RIFF",               # WAV
    b"\x00\x00\x00",       # MP4/M4A (ftyp box)
    b"ftyp",               # MP4/M4A
]


def _validate_audio(file_obj):
    pos = file_obj.tell()
    header = file_obj.read(16)
    file_obj.seek(pos)
    if len(header) < 3:
        return False
    for sig in AUDIO_SIGNATURES:
        if sig in header[:16]:
            return True
    return False


# ---------------------------------------------------------------------------
# Employee-facing: Estimate capture page
# ---------------------------------------------------------------------------

@estimates_bp.route("/estimate")
def estimate_capture():
    token_str = request.args.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data or not token_data["is_active"]:
        return render_template("errors/invalid_token.html"), 404

    h = _helpers()
    employee = h._require_employee_session(token_str)
    if not employee:
        return redirect(url_for("company_home", token_str=token_str))

    jobs = database.get_jobs_by_token(token_str, active_only=True)

    return render_template(
        "estimate_capture.html",
        token=token_data,
        employee=employee,
        jobs=jobs,
    )


# ---------------------------------------------------------------------------
# Employee-facing: Upload audio
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/upload", methods=["POST"])
def api_estimate_upload():
    h = _helpers()

    token_str = request.form.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data or not token_data["is_active"]:
        return jsonify({"error": "Invalid token"}), 403

    employee = h._require_employee_session(token_str)
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    job_id = request.form.get("job_id", type=int)
    if not job_id:
        return jsonify({"error": "Job is required"}), 400

    title = request.form.get("title", "").strip()

    audio = request.files.get("audio")
    audio_filename = ""

    if audio and audio.filename:
        if not _validate_audio(audio):
            return jsonify({"error": "Invalid audio file"}), 400

        # Save audio file
        now = datetime.now()
        timestamp_str = now.strftime("%Y%m%d_%H%M%S")
        unique_id = uuid.uuid4().hex[:8]

        audio_ext = "webm"
        if audio.filename:
            ext = audio.filename.rsplit(".", 1)[-1].lower() if "." in audio.filename else ""
            if ext in ("mp4", "m4a", "ogg", "wav", "webm"):
                audio_ext = ext
        audio_filename = f"estimate_{timestamp_str}_{unique_id}.{audio_ext}"

        folder = config.RECEIPTS_DIR / token_str / "estimates"
        folder.mkdir(parents=True, exist_ok=True)
        audio_path = folder / audio_filename
        audio.save(str(audio_path))

    # If no audio, mark as complete immediately (no transcription needed)
    status = "processing" if audio_filename else "complete"

    estimate = database.create_estimate(
        job_id=job_id,
        token_str=token_str,
        title=title,
        audio_file=audio_filename,
        created_by=employee["id"],
        status=status,
    )

    # Save customer info if provided
    customer_company_name = request.form.get("customer_company_name", "").strip()
    customer_name = request.form.get("customer_name", "").strip()
    customer_phone = request.form.get("customer_phone", "").strip()
    customer_email = request.form.get("customer_email", "").strip()
    if customer_company_name or customer_name or customer_phone or customer_email:
        database.update_estimate(
            estimate["id"],
            customer_company_name=customer_company_name,
            customer_name=customer_name,
            customer_phone=customer_phone,
            customer_email=customer_email,
        )

    return jsonify({"ok": True, "estimate_id": estimate["id"], "status": status})


# ---------------------------------------------------------------------------
# Employee-facing: Poll status
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/status/<int:estimate_id>")
def api_estimate_status(estimate_id):
    token_str = request.args.get("token", "")
    h = _helpers()
    employee = h._require_employee_session(token_str)
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    est = database.get_estimate(estimate_id)
    if not est or est["token"] != token_str:
        return jsonify({"error": "Not found"}), 404

    return jsonify({
        "status": est["status"],
        "transcription": est["transcription"] if est["status"] == "complete" else "",
    })


# ---------------------------------------------------------------------------
# Employee-facing: Quick add job
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/add-job", methods=["POST"])
def api_estimate_add_job():
    h = _helpers()
    token_str = request.form.get("token", "")
    employee = h._require_employee_session(token_str)
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    job_name = request.form.get("job_name", "").strip()
    job_address = request.form.get("job_address", "").strip()
    if not job_name:
        return jsonify({"error": "Job name is required"}), 400

    lat = request.form.get("latitude", type=float)
    lng = request.form.get("longitude", type=float)
    database.create_job(job_name, job_address, lat, lng, token_str)
    # Fetch the newly created job (last one by name+token)
    jobs = database.get_jobs_by_token(token_str)
    new_job = next((j for j in reversed(jobs) if j["job_name"] == job_name), None)
    if not new_job:
        return jsonify({"error": "Job creation failed"}), 500
    return jsonify({"ok": True, "job": {"id": new_job["id"], "job_name": new_job["job_name"]}})


# ---------------------------------------------------------------------------
# Employee-facing: Geocode address
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/geocode")
def api_estimate_geocode():
    """Geocode an address — accessible to authenticated employees."""
    h = _helpers()
    token_str = request.args.get("token", "")
    employee = h._require_employee_session(token_str)
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    address = request.args.get("address", "")
    if not address:
        return jsonify({"error": "Address required"}), 400

    import json
    import urllib.parse
    import urllib.request
    encoded = urllib.parse.quote(address)
    url = f"https://nominatim.openstreetmap.org/search?format=json&q={encoded}&limit=1"
    req = urllib.request.Request(url, headers={"User-Agent": "BDB-Tools/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            if data:
                return jsonify({
                    "lat": float(data[0]["lat"]),
                    "lng": float(data[0]["lon"]),
                })
            return jsonify({"error": "Address not found"}), 404
    except Exception:
        return jsonify({"error": "Geocoding service unavailable"}), 500


# ---------------------------------------------------------------------------
# Employee-facing: My Estimates list
# ---------------------------------------------------------------------------

@estimates_bp.route("/my-estimates")
def my_estimates():
    token_str = request.args.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data or not token_data["is_active"]:
        return render_template("errors/invalid_token.html"), 404

    h = _helpers()
    employee = h._require_employee_session(token_str)
    if not employee:
        return redirect(url_for("company_home", token_str=token_str))

    estimates = database.get_estimates_by_token(token_str)
    return render_template(
        "my_estimates.html",
        token=token_data,
        employee=employee,
        estimates=estimates,
    )


# ---------------------------------------------------------------------------
# Employee-facing: View/Edit single estimate
# ---------------------------------------------------------------------------

@estimates_bp.route("/my-estimates/<int:estimate_id>")
def my_estimate_detail(estimate_id):
    token_str = request.args.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data or not token_data["is_active"]:
        return render_template("errors/invalid_token.html"), 404

    h = _helpers()
    employee = h._require_employee_session(token_str)
    if not employee:
        return redirect(url_for("company_home", token_str=token_str))

    est = database.get_estimate(estimate_id)
    if not est or est["token"] != token_str:
        return render_template("errors/404.html"), 404

    job = database.get_job(est["job_id"])
    photos = database.get_all_job_photos_for_job(est["job_id"])
    items = database.get_estimate_items(estimate_id)
    products_services = database.get_products_services_by_token(token_str, active_only=True)

    _margin_target, _markup_required = _compute_finance_targets(token_str, token_data)

    return render_template(
        "my_estimate_detail.html",
        token=token_data,
        employee=employee,
        estimate=est,
        job=job,
        photos=photos,
        items=items,
        products_services=products_services,
        margin_target=_margin_target,
        markup_required=_markup_required,
    )


# ---------------------------------------------------------------------------
# Employee-facing: Update estimate (transcription/notes)
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/<int:estimate_id>/update", methods=["POST"])
def api_estimate_update(estimate_id):
    h = _helpers()
    token_str = request.form.get("token", "")
    employee = h._require_employee_session(token_str)
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    est = database.get_estimate(estimate_id)
    if not est or est["token"] != token_str:
        return jsonify({"error": "Not found"}), 404

    updates = {}
    for field in ("transcription", "notes", "title", "customer_company_name", "customer_name", "customer_phone", "customer_email"):
        if field in request.form:
            updates[field] = request.form[field].strip()
    for num_field in ("sales_tax_rate", "est_materials_cost", "est_labor_cost"):
        if num_field in request.form:
            try:
                updates[num_field] = max(0.0, float(request.form[num_field]))
            except (ValueError, TypeError):
                pass

    if updates:
        database.update_estimate(estimate_id, **updates)

    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Employee-facing: Append additional voice memo to existing estimate
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/<int:estimate_id>/add-audio", methods=["POST"])
def api_estimate_add_audio(estimate_id):
    import os
    import time as _time

    h = _helpers()
    token_str = request.form.get("token", "")
    employee = h._require_employee_session(token_str)
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    est = database.get_estimate(estimate_id)
    if not est or est["token"] != token_str:
        return jsonify({"error": "Not found"}), 404

    audio_file = request.files.get("audio")
    if not audio_file or not audio_file.filename:
        return jsonify({"error": "No audio file provided"}), 400

    if not _validate_audio(audio_file):
        return jsonify({"error": "Invalid audio file"}), 400

    audio_ext = "webm"
    if audio_file.filename and "." in audio_file.filename:
        ext = audio_file.filename.rsplit(".", 1)[-1].lower()
        if ext in ("mp4", "m4a", "ogg", "wav", "webm"):
            audio_ext = ext

    filename = f"append_{estimate_id}_{int(_time.time())}.{audio_ext}"
    folder = config.RECEIPTS_DIR / token_str / "estimates"
    folder.mkdir(parents=True, exist_ok=True)
    save_path = folder / filename
    audio_file.save(str(save_path))

    database.update_estimate(estimate_id, append_audio_file=str(save_path), status="appending")
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Employee-facing: Line item CRUD
# ---------------------------------------------------------------------------

@estimates_bp.route("/api/estimate/<int:estimate_id>/items/create", methods=["POST"])
def api_employee_item_create(estimate_id):
    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h = _helpers()
    employee = h._require_employee_session(est["token"])
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    data = request.get_json() or {}
    description = (data.get("description") or "").strip()
    if not description:
        return jsonify({"error": "Description required"}), 400

    try:
        quantity = max(0.0, float(data.get("quantity", 1) or 1))
        unit_price = max(0.0, float(data.get("unit_price", 0) or 0))
    except (ValueError, TypeError):
        quantity, unit_price = 1.0, 0.0
    try:
        unit_cost = max(0.0, float(data.get("unit_cost", 0) or 0))
    except (ValueError, TypeError):
        unit_cost = 0.0
    total = round(quantity * unit_price, 2)
    taxable = 1 if data.get("taxable") else 0
    item_type = data.get("item_type", "product")
    if item_type not in ("product", "service"):
        item_type = "product"

    item = database.create_estimate_item(
        estimate_id, est["token"], description, quantity, unit_price, total, taxable, 0, item_type,
        unit_cost=unit_cost
    )
    return jsonify({"ok": True, "item": item})


@estimates_bp.route("/api/estimate/items/<int:item_id>/update", methods=["POST"])
def api_employee_item_update(item_id):
    conn = database.get_db()
    item = conn.execute("SELECT * FROM estimate_items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not item:
        return jsonify({"error": "Not found"}), 404
    h = _helpers()
    employee = h._require_employee_session(item["token"])
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    data = request.get_json() or {}
    updates = {}
    if "description" in data:
        updates["description"] = (data["description"] or "").strip()
    for f in ("quantity", "unit_price", "unit_cost"):
        if f in data:
            try:
                updates[f] = max(0.0, float(data[f]))
            except (ValueError, TypeError):
                pass
    if "taxable" in data:
        updates["taxable"] = 1 if data["taxable"] else 0
    if "item_type" in data:
        val = data["item_type"]
        if val in ("product", "service"):
            updates["item_type"] = val

    # Always recompute total from qty × unit_price server-side
    qty = updates.get("quantity", item["quantity"] or 0)
    price = updates.get("unit_price", item["unit_price"] or 0)
    updates["total"] = round(float(qty) * float(price), 2)

    if updates:
        database.update_estimate_item(item_id, **updates)
    return jsonify({"ok": True})


@estimates_bp.route("/api/estimate/items/<int:item_id>/delete", methods=["POST"])
def api_employee_item_delete(item_id):
    conn = database.get_db()
    item = conn.execute("SELECT * FROM estimate_items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not item:
        return jsonify({"error": "Not found"}), 404
    h = _helpers()
    employee = h._require_employee_session(item["token"])
    if not employee:
        return jsonify({"error": "Not authenticated"}), 401

    database.delete_estimate_item(item_id)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Admin: Estimates list
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates")
def admin_estimates():
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    token_str = request.args.get("token", "")
    tokens = database.get_all_tokens()
    selected_token = None

    if token_str:
        selected_token = database.get_token(token_str)
    elif tokens:
        selected_token = tokens[0]
        token_str = selected_token["token"]

    estimates = []
    filter_job = None
    back_customer_id = request.args.get("back_customer", type=int)
    job_id_filter = request.args.get("job_id", type=int)
    if token_str:
        search = request.args.get("search", "")
        estimates = database.get_estimates_by_token(
            token_str, search=search or None, job_id=job_id_filter or None
        )
        if job_id_filter:
            filter_job = database.get_job(job_id_filter)

    return render_template(
        "admin/estimates.html",
        tokens=tokens,
        selected_token=selected_token,
        estimates=estimates,
        search=request.args.get("search", ""),
        filter_job=filter_job,
        back_customer_id=back_customer_id,
    )


# ---------------------------------------------------------------------------
# Admin: Create blank estimate
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/create", methods=["POST"])
def admin_estimate_create():
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    token_str = request.form.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data:
        return redirect(url_for("estimates.admin_estimates"))
    h._verify_token_access(token_str)

    # Prefer explicit job_id from form (set when creating in job/customer context)
    job_id = request.form.get("job_id", type=int)
    if not job_id:
        jobs = database.get_jobs_by_token(token_str, active_only=True)
        job_id = jobs[0]["id"] if jobs else None

    est = database.create_estimate(
        job_id=job_id,
        token_str=token_str,
        title="",
        audio_file="",
        status="complete",
        created_by=getattr(h.current_user, "id", None),
    )
    customer_id = request.form.get("customer_id", type=int)
    if customer_id:
        database.link_estimate_to_customer(est["id"], customer_id, token_str)
    return redirect(f"/admin/estimates/{est['id']}?token={token_str}")


# ---------------------------------------------------------------------------
# Admin: Estimate detail
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>")
def admin_estimate_detail(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    est = database.get_estimate(estimate_id)
    if not est:
        return render_template("errors/404.html"), 404
    h._verify_token_access(est["token"])

    job = database.get_job(est["job_id"])
    photos = database.get_all_job_photos_for_job(est["job_id"])
    token_data = database.get_token(est["token"])
    items = database.get_estimate_items(estimate_id)
    snippets = database.get_message_snippets_by_token(est["token"], active_only=True)
    products_services = database.get_products_services_by_token(est["token"], active_only=True)
    job_labor = database.get_job_labor_total(est["job_id"], est["token"])
    is_project = est.get("approval_status", "pending") not in ("pending", "declined")
    all_jobs = database.get_jobs_by_token(est["token"])

    _margin_target, _markup_required = _compute_finance_targets(est["token"], token_data)
    customers = database.get_customers_by_token(est["token"], active_only=True)

    # Task templates pool for this project
    linked_templates = database.get_templates_for_estimate(estimate_id, est["token"])
    all_templates = database.get_task_templates(est["token"])
    available_templates = [t for t in all_templates
                           if t["id"] not in {lt["id"] for lt in linked_templates}]
    project_tasks = database.get_project_tasks_by_estimate(estimate_id, est["token"])

    return render_template(
        "admin/estimate_detail.html",
        estimate=est,
        job=job,
        photos=photos,
        selected_token=token_data,
        items=items,
        snippets=snippets,
        products_services=products_services,
        job_labor=job_labor,
        is_project=is_project,
        all_jobs=all_jobs,
        margin_target=_margin_target,
        markup_required=_markup_required,
        customers=customers,
        linked_templates=linked_templates,
        available_templates=available_templates,
        project_tasks=project_tasks,
    )


# ---------------------------------------------------------------------------
# Admin: Estimate — Task Template Pool (apply/remove)
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/templates/apply", methods=["POST"])
def admin_estimate_apply_template(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))
    est = database.get_estimate(estimate_id)
    if not est:
        abort(404)
    h._verify_token_access(est["token"])
    _require_admin()
    template_id = request.form.get("template_id", type=int)
    if template_id:
        database.apply_template_to_estimate(estimate_id, template_id, est["token"])
        flash("Template added to project.", "success")
    return redirect(url_for("estimates.admin_estimate_detail", estimate_id=estimate_id))


@estimates_bp.route("/admin/estimates/<int:estimate_id>/templates/<int:template_id>/remove", methods=["POST"])
def admin_estimate_remove_template(estimate_id, template_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))
    est = database.get_estimate(estimate_id)
    if not est:
        abort(404)
    h._verify_token_access(est["token"])
    _require_admin()
    database.remove_template_from_estimate(estimate_id, template_id, est["token"])
    flash("Template removed from project.", "success")
    return redirect(url_for("estimates.admin_estimate_detail", estimate_id=estimate_id))


# ---------------------------------------------------------------------------
# Admin: Estimate — Project Specific Tasks CRUD
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/tasks/create", methods=["POST"])
def admin_estimate_task_create(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))
    est = database.get_estimate(estimate_id)
    if not est:
        abort(404)
    h._verify_token_access(est["token"])
    _require_admin()
    name = request.form.get("name", "").strip()
    if name:
        database.create_project_task(estimate_id, est["job_id"], name, est["token"])
        flash("Task added.", "success")
    return redirect(url_for("estimates.admin_estimate_detail", estimate_id=estimate_id))


@estimates_bp.route("/admin/estimates/<int:estimate_id>/tasks/create-json", methods=["POST"])
def admin_estimate_task_create_json(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Not authenticated"}), 401
    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(est["token"])
    _require_admin()
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    task_id = database.create_project_task(estimate_id, est["job_id"], name, est["token"])
    return jsonify({"ok": True, "id": task_id, "name": name})


@estimates_bp.route("/admin/estimates/<int:estimate_id>/tasks/<int:task_id>/delete", methods=["POST"])
def admin_estimate_task_delete(estimate_id, task_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))
    est = database.get_estimate(estimate_id)
    if not est:
        abort(404)
    h._verify_token_access(est["token"])
    _require_admin()
    database.delete_project_task(task_id, est["token"])
    flash("Task removed.", "success")
    return redirect(url_for("estimates.admin_estimate_detail", estimate_id=estimate_id))


# ---------------------------------------------------------------------------
# Admin: Jobs list (JSON for AJAX refresh)
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/jobs.json")
def admin_estimate_jobs_json():
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify([])
    token_str = request.args.get("token", "")
    if not token_str:
        return jsonify([])
    h._verify_token_access(token_str)
    jobs = database.get_jobs_by_token(token_str)
    return jsonify([{"id": j["id"], "job_name": j["job_name"], "job_address": j.get("job_address", "")} for j in jobs])


# ---------------------------------------------------------------------------
# Admin: Update estimate
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/update", methods=["POST"])
def admin_estimate_update(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(est["token"])

    if request.is_json:
        data = request.get_json()
    else:
        data = request.form

    updates = {}
    for field in ("transcription", "notes", "title", "customer_company_name", "customer_name", "customer_phone", "customer_email",
                   "estimate_number", "date_accepted", "expected_completion", "customer_message", "project_name"):
        if field in data:
            val = data[field]
            updates[field] = val.strip() if isinstance(val, str) else val
    if "approval_status" in data:
        val = data["approval_status"].strip() if isinstance(data["approval_status"], str) else data["approval_status"]
        if val in ("pending", "accepted", "in_progress", "completed", "declined"):
            updates["approval_status"] = val
    for num_field in ("estimate_value", "est_materials_cost", "est_labor_cost",
                       "actual_materials_cost", "actual_collected",
                       "est_labor_hours", "sales_tax_rate",
                       "completion_pct", "client_budget"):
        if num_field in data:
            try:
                val = float(data[num_field])
                if 0 <= val <= 999999999:
                    updates[num_field] = val
            except (ValueError, TypeError):
                pass

    # Validate estimate_number uniqueness within this company
    if "estimate_number" in updates and updates["estimate_number"]:
        if database.is_estimate_number_taken(est["token"], updates["estimate_number"], exclude_id=estimate_id):
            return jsonify({"error": "That number is already in use by another estimate/project."}), 400

    # Handle customer_id BEFORE calling update_estimate so it is included in the DB write
    new_customer_id_response = None
    if "customer_id" in data:
        raw_cid = data["customer_id"]
        try:
            cid_int = int(raw_cid) if raw_cid else None
            if cid_int == -1:
                # Create new customer from the estimate's contact fields
                cname = (updates.get("customer_name") or "").strip()
                if cname:
                    new_cid = database.create_customer(
                        company_name=(updates.get("customer_company_name") or "").strip(),
                        customer_name=cname,
                        phone=(updates.get("customer_phone") or "").strip(),
                        email=(updates.get("customer_email") or "").strip(),
                        notes="",
                        token_str=est["token"],
                    )
                    updates["customer_id"] = new_cid
                    new_customer_id_response = new_cid
                # else: no name provided — ignore the -1 sentinel
            else:
                updates["customer_id"] = cid_int
        except (ValueError, TypeError):
            pass

    if updates:
        database.update_estimate(estimate_id, **updates)

    # Auto-link parent job when customer is assigned
    final_customer_id = updates.get("customer_id")
    if final_customer_id and est.get("job_id"):
        database.link_job_to_customer(est["job_id"], final_customer_id, est["token"])

    # Handle job reassignment
    job_id_str = data.get("job_id")
    if job_id_str:
        try:
            new_job_id = int(job_id_str)
            if new_job_id != est.get("job_id"):
                database.update_estimate(estimate_id, job_id=new_job_id)
        except (ValueError, TypeError):
            pass

    if request.is_json:
        resp = {"ok": True}
        if new_customer_id_response:
            resp["new_customer_id"] = new_customer_id_response
            c = database.get_customer(new_customer_id_response)
            resp["new_company_name"] = c["company_name"] if c else ""
        return jsonify(resp)
    return redirect(f"/admin/estimates/{estimate_id}?token={est['token']}")


# ---------------------------------------------------------------------------
# Admin: Delete estimate
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/delete", methods=["POST"])
def admin_estimate_delete(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(est["token"])

    token = est["token"]
    database.delete_estimate(estimate_id)

    if request.is_json:
        return jsonify({"ok": True})
    return redirect(f"/admin/estimates?token={token}")


# ---------------------------------------------------------------------------
# Admin: Estimate PDF report
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/report/pdf")
def admin_estimate_pdf(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    est = database.get_estimate(estimate_id)
    if not est:
        return render_template("errors/404.html"), 404
    h._verify_token_access(est["token"])

    job = database.get_job(est["job_id"])
    job_name = job["job_name"] if job else "Unknown Job"
    photos = database.get_all_job_photos_for_job(est["job_id"])
    tasks = database.get_job_tasks(est["job_id"])
    token_data = database.get_token(est["token"])
    company_name = token_data["company_name"] if token_data else ""

    # Filter photos by selected IDs if provided
    if "photos" in request.args:
        try:
            selected_ids = {int(x) for x in request.args.get("photos", "").split(",") if x.strip()}
            photos = [p for p in photos if p["id"] in selected_ids]
        except ValueError:
            photos = []

    job_labor = database.get_job_labor_total(est["job_id"], est["token"])
    est = dict(est)
    est["actual_labor_hours"] = job_labor["total_hours"]
    est["actual_labor_cost"] = job_labor["total_cost"]

    import pdf_generator
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_generator.generate_estimate_pdf(
            output_path=tmp.name,
            estimate=est,
            job_name=job_name,
            photos=photos,
            tasks=tasks,
            company_name=company_name,
        )
        return send_file(
            tmp.name,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{'project' if est.get('approval_status', 'pending') not in ('pending', 'declined') else 'estimate'}_{estimate_id}_{job_name[:30]}.pdf",
        )


# ---------------------------------------------------------------------------
# Admin: Estimate Excel report
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/report/xlsx")
def admin_estimate_xlsx(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    est = database.get_estimate(estimate_id)
    if not est:
        return render_template("errors/404.html"), 404
    h._verify_token_access(est["token"])

    job = database.get_job(est["job_id"])
    job_name = job["job_name"] if job else "Unknown Job"
    tasks = database.get_job_tasks(est["job_id"])
    items = database.get_estimate_items(estimate_id)
    token_data = database.get_token(est["token"])
    company_name = token_data["company_name"] if token_data else ""

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, numbers
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()

    # --- Styles ---
    wrap_align = Alignment(wrap_text=True, vertical="top")
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    label_font = Font(bold=False, size=10)
    value_font = Font(bold=False, size=10)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    def write_section_header(ws, row, label, span=3):
        c = ws.cell(row=row, column=1, value=label)
        c.font = section_font
        c.fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        for col_idx in range(2, span + 1):
            ws.cell(row=row, column=col_idx).fill = section_fill
        return row + 1

    def write_row(ws, row, label, value, col=2, fmt=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = label_font
        lc.alignment = wrap_align
        c = ws.cell(row=row, column=col, value=value)
        c.font = value_font
        c.alignment = wrap_align
        if fmt:
            c.number_format = fmt
        return row + 1

    def write_variance_row(ws, row, label, value, fmt=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = label_font
        lc.alignment = wrap_align
        c = ws.cell(row=row, column=2, value=value)
        c.font = value_font
        c.alignment = wrap_align
        if fmt:
            c.number_format = fmt
        if isinstance(value, (int, float)):
            c.fill = green_fill if value >= 0 else red_fill
        return row + 1

    def write_text_block(ws, row, label, text, span=6):
        """Write a section header + large text block with wrap and fixed row height."""
        row = write_section_header(ws, row, label, span=span)
        c = ws.cell(row=row, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        ws.row_dimensions[row].height = 50
        return row + 2

    # ===== Tab 1: Financial Summary =====
    ws1 = wb.active
    ws1.title = "Financial Summary"

    # Compute financial values
    job_labor = database.get_job_labor_total(est["job_id"], est["token"])
    est_value = est.get("estimate_value") or 0
    est_mat = est.get("est_materials_cost") or 0
    est_lab = est.get("est_labor_cost") or 0
    est_hrs = est.get("est_labor_hours") or 0
    act_mat = est.get("actual_materials_cost") or 0
    act_lab = job_labor["total_cost"]
    act_hrs = job_labor["total_hours"]
    act_collected = est.get("actual_collected") or 0
    completion_pct = est.get("completion_pct") or 0

    est_cost = est_mat + est_lab
    est_profit = est_value - est_cost
    est_margin = (est_profit / est_value) if est_value else 0
    est_markup = (est_profit / est_cost) if est_cost else 0

    pct_factor = completion_pct / 100
    wip_ev = est_value * pct_factor
    wip_mat = est_mat * pct_factor
    wip_lab = est_lab * pct_factor
    wip_hrs = est_hrs * pct_factor
    wip_cost = wip_mat + wip_lab
    wip_profit = wip_ev - wip_cost
    wip_margin = (wip_profit / wip_ev) if wip_ev else 0
    wip_markup = (wip_profit / wip_cost) if wip_cost else 0

    act_cost = act_mat + act_lab
    act_profit = act_collected - act_cost
    act_margin = (act_profit / act_collected) if act_collected else 0
    act_markup = (act_profit / act_cost) if act_cost else 0

    row = 1

    # Header
    doc_label = "Project" if est.get("approval_status", "pending") not in ("pending", "declined") else "Estimate"
    doc_num = est.get("estimate_number") or str(est["id"])
    c = ws1.cell(row=row, column=1, value=f"Financial Summary — {doc_label} #{doc_num}")
    c.font = header_font
    row += 1

    row = write_section_header(ws1, row, f"Company & {doc_label}")
    row = write_row(ws1, row, "Company", company_name)
    row = write_row(ws1, row, f"{doc_label} #", doc_num)
    row = write_row(ws1, row, "Created Date", est["created_at"][:10])
    status_labels = {"pending": "Pending", "accepted": "Accepted", "in_progress": "In Progress", "completed": "Completed", "declined": "Declined"}
    row = write_row(ws1, row, "Approval Status", status_labels.get(est.get("approval_status", ""), est.get("approval_status", "")))
    row = write_row(ws1, row, "Completion %", completion_pct / 100, fmt=pct_fmt)
    row += 1

    row = write_section_header(ws1, row, "Customer")
    row = write_row(ws1, row, "Name", est.get("customer_name") or "")
    row = write_row(ws1, row, "Phone", est.get("customer_phone") or "")
    row = write_row(ws1, row, "Email", est.get("customer_email") or "")
    row += 1

    row = write_section_header(ws1, row, "Job")
    row = write_row(ws1, row, "Job Name", job_name)
    row = write_row(ws1, row, "Address", job.get("job_address", "") if job else "")
    row += 1

    row = write_section_header(ws1, row, "Estimate Overview")
    row = write_row(ws1, row, "Estimate Value", est_value, fmt=currency_fmt)
    row = write_row(ws1, row, "Date Accepted", est.get("date_accepted") or "")
    row = write_row(ws1, row, "Expected Completion", est.get("expected_completion") or "")
    row += 1

    row = write_section_header(ws1, row, "Estimated Costs")
    row = write_row(ws1, row, "Materials", est_mat, fmt=currency_fmt)
    row = write_row(ws1, row, "Labor", est_lab, fmt=currency_fmt)
    row = write_row(ws1, row, "Labor Hours", est_hrs)
    row = write_row(ws1, row, "Margin %", est_margin, fmt=pct_fmt)
    row = write_row(ws1, row, "Markup %", est_markup, fmt=pct_fmt)
    row = write_row(ws1, row, "Gross Profit", est_profit, fmt=currency_fmt)
    row += 1

    row = write_section_header(ws1, row, f"Work in Progress ({completion_pct:.0f}% Complete)")
    row = write_row(ws1, row, "Materials", wip_mat, fmt=currency_fmt)
    row = write_row(ws1, row, "Labor", wip_lab, fmt=currency_fmt)
    row = write_row(ws1, row, "Labor Hours", wip_hrs)
    row = write_row(ws1, row, "Margin %", wip_margin, fmt=pct_fmt)
    row = write_row(ws1, row, "Markup %", wip_markup, fmt=pct_fmt)
    row = write_row(ws1, row, "Gross Profit", wip_profit, fmt=currency_fmt)
    row += 1

    row = write_section_header(ws1, row, "Actual Costs")
    row = write_row(ws1, row, "Materials", act_mat, fmt=currency_fmt)
    row = write_row(ws1, row, "Labor", act_lab, fmt=currency_fmt)
    row = write_row(ws1, row, "Labor Hours", act_hrs)
    row = write_row(ws1, row, "Margin %", act_margin, fmt=pct_fmt)
    row = write_row(ws1, row, "Markup %", act_markup, fmt=pct_fmt)
    row = write_row(ws1, row, "Gross Profit", act_profit, fmt=currency_fmt)
    row += 1

    row = write_section_header(ws1, row, "Variance (Work in Progress vs Actual)")
    row = write_variance_row(ws1, row, "Materials", wip_mat - act_mat, fmt=currency_fmt)
    row = write_variance_row(ws1, row, "Labor", wip_lab - act_lab, fmt=currency_fmt)
    row = write_variance_row(ws1, row, "Hours", wip_hrs - act_hrs)
    row = write_variance_row(ws1, row, "Margin pts", (wip_margin - act_margin) * 100)
    row = write_variance_row(ws1, row, "Markup pts", (wip_markup - act_markup) * 100)
    row = write_variance_row(ws1, row, "Profit $", wip_profit - act_profit, fmt=currency_fmt)
    row += 1

    row = write_section_header(ws1, row, "Collection")
    row = write_row(ws1, row, "Estimate Value", est_value, fmt=currency_fmt)
    row = write_row(ws1, row, "Actual Collected", act_collected, fmt=currency_fmt)
    collection_diff = act_collected - est_value
    r = row
    row = write_variance_row(ws1, row, "Difference", collection_diff, fmt=currency_fmt)

    # Auto column widths
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    # Apply wrap to all Tab 1 cells
    for r in range(1, row + 1):
        for col_idx in range(1, 4):
            cell = ws1.cell(row=r, column=col_idx)
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ===== Tab 2: Estimate/Project Details =====
    ws2 = wb.create_sheet(f"{doc_label} Details")
    row = 1

    c = ws2.cell(row=row, column=1, value=f"{doc_label} Details — #{doc_num}")
    c.font = header_font
    row += 2

    # Line Items table
    row = write_section_header(ws2, row, "Line Items", span=6)
    # Table header
    headers = ["#", "Description", "Qty", "Unit Price", "Taxable", "Total"]
    table_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx, hdr in enumerate(headers, 1):
        c = ws2.cell(row=row, column=col_idx, value=hdr)
        c.font = Font(bold=True, size=10)
        c.fill = table_fill
    row += 1

    for i, item in enumerate(items, 1):
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=item.get("description", ""))
        ws2.cell(row=row, column=3, value=item.get("quantity", 0))
        c = ws2.cell(row=row, column=4, value=item.get("unit_price", 0))
        c.number_format = currency_fmt
        ws2.cell(row=row, column=5, value="Yes" if item.get("taxable") else "No")
        c = ws2.cell(row=row, column=6, value=item.get("total", 0))
        c.number_format = currency_fmt
        row += 1
    row += 1

    # Totals
    subtotal = sum(item.get("total", 0) for item in items)
    tax_rate = est.get("sales_tax_rate", 0) or 0
    taxable_total = sum(item.get("total", 0) for item in items if item.get("taxable"))
    sales_tax = taxable_total * (tax_rate / 100)
    grand_total = subtotal + sales_tax

    row = write_section_header(ws2, row, "Totals", span=6)
    row = write_row(ws2, row, "Subtotal", subtotal, fmt=currency_fmt)
    row = write_row(ws2, row, "Tax Rate", tax_rate / 100, fmt=pct_fmt)
    row = write_row(ws2, row, "Sales Tax", sales_tax, fmt=currency_fmt)
    ws2.cell(row=row, column=1, value="Grand Total").font = Font(bold=True, size=11)
    c = ws2.cell(row=row, column=2, value=grand_total)
    c.font = Font(bold=True, size=11)
    c.number_format = currency_fmt
    row += 2

    # Customer Message
    msg = (est.get("customer_message") or "").strip()
    if msg:
        row = write_text_block(ws2, row, "Customer Message", msg)

    # Transcription
    if est.get("transcription"):
        row = write_text_block(ws2, row, "Transcription", est["transcription"])

    # Notes
    if est.get("notes"):
        row = write_text_block(ws2, row, "Notes", est["notes"])

    # Tasks table
    if tasks:
        row = write_section_header(ws2, row, "Tasks", span=6)
        task_headers = ["#", "Task Name", "Source", "Active"]
        for col_idx, hdr in enumerate(task_headers, 1):
            c = ws2.cell(row=row, column=col_idx, value=hdr)
            c.font = Font(bold=True, size=10)
            c.fill = table_fill
        row += 1
        for i, t in enumerate(tasks, 1):
            ws2.cell(row=row, column=1, value=i)
            ws2.cell(row=row, column=2, value=t["name"])
            ws2.cell(row=row, column=3, value=t["source"])
            ws2.cell(row=row, column=4, value="Yes" if t["is_active"] else "No")
            row += 1

    # Auto column widths for Tab 2 (total ~70 for text block merges)
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 8
    ws2.column_dimensions["F"].width = 12
    # Set default row height and wrap for description cells
    for r in range(1, row + 1):
        for col_idx in range(1, 7):
            cell = ws2.cell(row=r, column=col_idx)
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"financial_report_{estimate_id}_{job_name[:30]}.xlsx",
    )


# ---------------------------------------------------------------------------
# Admin: Estimate line items API
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/items/create", methods=["POST"])
def admin_estimate_item_create(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(est["token"])

    data = request.get_json() if request.is_json else request.form
    description = (data.get("description") or "").strip()
    if not description:
        return jsonify({"error": "Description is required"}), 400

    try:
        quantity = float(data.get("quantity", 1))
        unit_price = float(data.get("unit_price", 0))
        total = float(data.get("total", 0))
        sort_order = int(data.get("sort_order", 0))
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid numeric value"}), 400
    try:
        unit_cost = max(0.0, float(data.get("unit_cost", 0) or 0))
    except (ValueError, TypeError):
        unit_cost = 0.0

    taxable = 1 if data.get("taxable") in (True, 1, "1", "true", "on") else 0
    item_type = data.get("item_type", "product")
    if item_type not in ("product", "service"):
        item_type = "product"

    item = database.create_estimate_item(
        estimate_id, est["token"], description, quantity, unit_price, total, taxable, sort_order, item_type,
        unit_cost=unit_cost
    )
    return jsonify({"ok": True, "item": item})


@estimates_bp.route("/admin/estimates/items/<int:item_id>/update", methods=["POST"])
def admin_estimate_item_update(item_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    conn = database.get_db()
    item = conn.execute("SELECT * FROM estimate_items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not item:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(item["token"])

    data = request.get_json() if request.is_json else request.form
    updates = {}

    if "description" in data:
        updates["description"] = (data["description"] or "").strip()
    for num_field in ("quantity", "unit_price", "unit_cost", "total"):
        if num_field in data:
            try:
                updates[num_field] = float(data[num_field])
            except (ValueError, TypeError):
                pass
    if "taxable" in data:
        updates["taxable"] = 1 if data["taxable"] in (True, 1, "1", "true", "on") else 0
    if "sort_order" in data:
        try:
            updates["sort_order"] = int(data["sort_order"])
        except (ValueError, TypeError):
            pass
    if "item_type" in data:
        val = data.get("item_type")
        if val in ("product", "service"):
            updates["item_type"] = val

    if updates:
        database.update_estimate_item(item_id, **updates)
    return jsonify({"ok": True})


@estimates_bp.route("/admin/estimates/items/<int:item_id>/delete", methods=["POST"])
def admin_estimate_item_delete(item_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    conn = database.get_db()
    item = conn.execute("SELECT * FROM estimate_items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not item:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(item["token"])

    database.delete_estimate_item(item_id)
    return jsonify({"ok": True})


@estimates_bp.route("/admin/estimates/<int:estimate_id>/items/save-product", methods=["POST"])
def admin_estimate_item_save_product(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(est["token"])

    data = request.get_json() if request.is_json else request.form
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    try:
        unit_price = float(data.get("unit_price", 0))
    except (ValueError, TypeError):
        unit_price = 0
    try:
        unit_cost = float(data.get("unit_cost", 0))
    except (ValueError, TypeError):
        unit_cost = 0
    item_type = data.get("item_type", "product")
    if item_type not in ("product", "service"):
        item_type = "product"

    ps = database.create_product_service(name, unit_price, est["token"], unit_cost=unit_cost, item_type=item_type)
    return jsonify({"ok": True, "product": ps})


# ---------------------------------------------------------------------------
# Employee-facing: Client estimate PDF (inline view)
# ---------------------------------------------------------------------------

@estimates_bp.route("/my-estimates/<int:estimate_id>/client-pdf")
def my_estimate_client_pdf(estimate_id):
    token_str = request.args.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data or not token_data["is_active"]:
        return render_template("errors/invalid_token.html"), 404

    h = _helpers()
    employee = h._require_employee_session(token_str)
    if not employee:
        return redirect(url_for("company_home", token_str=token_str))

    est = database.get_estimate(estimate_id)
    if not est or est["token"] != token_str:
        return render_template("errors/404.html"), 404

    job = database.get_job(est["job_id"])
    items = database.get_estimate_items(estimate_id)
    photos = database.get_all_job_photos_for_job(est["job_id"])

    if "photos" in request.args:
        try:
            selected_ids = {int(x) for x in request.args.get("photos", "").split(",") if x.strip()}
            photos = [p for p in photos if p["id"] in selected_ids]
        except ValueError:
            photos = []

    import pdf_generator
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_generator.generate_client_estimate_pdf(
            output_path=tmp.name,
            estimate=est,
            job=job,
            items=items,
            token_data=token_data,
            photos=photos,
        )
        _cust = re.sub(r"[^a-zA-Z0-9 \-]", "", (est.get("customer_name") or "").strip())[:25]
        _jname = re.sub(r"[^a-zA-Z0-9 \-]", "", (job["job_name"] if job else "estimate").strip())[:30]
        _date = (est.get("created_at") or "")[:10]
        fname = "-".join(p for p in [_cust, _jname, _date] if p) + ".pdf"
        # ?share=1 returns octet-stream so iOS Safari's fetch() doesn't intercept it as a PDF viewer
        if request.args.get("share"):
            return send_file(tmp.name, mimetype="application/octet-stream", as_attachment=True, download_name=fname)
        return send_file(tmp.name, mimetype="application/pdf", as_attachment=False, download_name=fname)


# ---------------------------------------------------------------------------
# Admin: Scope of Work PDF
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/report/scope-pdf")
def admin_estimate_scope_pdf(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    est = database.get_estimate(estimate_id)
    if not est:
        return render_template("errors/404.html"), 404
    h._verify_token_access(est["token"])

    job = database.get_job(est["job_id"])
    job_name = job["job_name"] if job else "Unknown Job"
    items = database.get_estimate_items(estimate_id)
    token_data = database.get_token(est["token"])
    company_name = token_data["company_name"] if token_data else ""
    photos = database.get_all_job_photos_for_job(est["job_id"])

    if "photos" in request.args:
        try:
            selected_ids = {int(x) for x in request.args.get("photos", "").split(",") if x.strip()}
            photos = [p for p in photos if p["id"] in selected_ids]
        except ValueError:
            photos = []

    import pdf_generator
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_generator.generate_scope_of_work_pdf(
            output_path=tmp.name,
            estimate=est,
            job=job,
            items=items,
            company_name=company_name,
            photos=photos,
        )
        _cust = re.sub(r"[^a-zA-Z0-9 \-]", "", (est.get("customer_name") or "").strip())[:25]
        _jname = re.sub(r"[^a-zA-Z0-9 \-]", "", job_name.strip())[:30]
        _date = (est.get("created_at") or "")[:10]
        _fname = "-".join(p for p in [_cust, _jname, _date] if p) + ".pdf"
        return send_file(
            tmp.name,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=_fname,
        )


# ---------------------------------------------------------------------------
# Employee-facing: Scope of Work PDF
# ---------------------------------------------------------------------------

@estimates_bp.route("/my-estimates/<int:estimate_id>/scope-pdf")
def my_estimate_scope_pdf(estimate_id):
    token_str = request.args.get("token", "")
    token_data = database.get_token(token_str)
    if not token_data or not token_data["is_active"]:
        return render_template("errors/invalid_token.html"), 404

    h = _helpers()
    employee = h._require_employee_session(token_str)
    if not employee:
        return redirect(url_for("company_home", token_str=token_str))

    est = database.get_estimate(estimate_id)
    if not est or est["token"] != token_str:
        return render_template("errors/404.html"), 404

    job = database.get_job(est["job_id"])
    job_name = job["job_name"] if job else "Unknown Job"
    items = database.get_estimate_items(estimate_id)
    company_name = token_data["company_name"] if token_data else ""
    photos = database.get_all_job_photos_for_job(est["job_id"])

    if "photos" in request.args:
        try:
            selected_ids = {int(x) for x in request.args.get("photos", "").split(",") if x.strip()}
            photos = [p for p in photos if p["id"] in selected_ids]
        except ValueError:
            photos = []

    import pdf_generator
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_generator.generate_scope_of_work_pdf(
            output_path=tmp.name,
            estimate=est,
            job=job,
            items=items,
            company_name=company_name,
            photos=photos,
        )
        _cust = re.sub(r"[^a-zA-Z0-9 \-]", "", (est.get("customer_name") or "").strip())[:25]
        _jname = re.sub(r"[^a-zA-Z0-9 \-]", "", job_name.strip())[:30]
        _date = (est.get("created_at") or "")[:10]
        _fname = "-".join(p for p in [_cust, _jname, _date] if p) + ".pdf"
        return send_file(
            tmp.name,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=_fname,
        )


# ---------------------------------------------------------------------------
# Admin: Send Scope PDF to Job Folder
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/send-to-job-folder", methods=["POST"])
def admin_estimate_send_to_job_folder(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    est = database.get_estimate(estimate_id)
    if not est:
        return jsonify({"error": "Not found"}), 404
    h._verify_token_access(est["token"])

    job = database.get_job(est["job_id"])
    if not job:
        return jsonify({"error": "Job not found"}), 404

    items = database.get_estimate_items(estimate_id)
    token_data = database.get_token(est["token"])
    company_name = token_data["company_name"] if token_data else ""
    token_str = est["token"]

    import pdf_generator
    import tempfile

    # Generate the scope PDF to a temp file
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_generator.generate_scope_of_work_pdf(
            output_path=tmp.name,
            estimate=est,
            job=job,
            items=items,
            company_name=company_name,
        )
        tmp_path = tmp.name

    # Build destination path using same patterns as job_photos upload
    def _sanitize(name):
        sanitized = re.sub(r"[^a-zA-Z0-9]+", "-", name)
        return sanitized.strip("-")

    now = datetime.now()
    iso = now.isocalendar()
    week_folder = f"{iso[0]}-W{iso[1]:02d}"
    safe_job_name = _sanitize(job["job_name"])
    pdf_filename = f"scope_of_work_{estimate_id}.pdf"

    dest_dir = config.JOB_PHOTOS_DIR / token_str / safe_job_name / week_folder
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / pdf_filename

    import shutil
    shutil.copy2(tmp_path, str(dest_path))

    import os
    os.unlink(tmp_path)

    # Store relative path in DB (from JOB_PHOTOS_DIR) — skip if already exists
    rel_path = f"{token_str}/{safe_job_name}/{week_folder}/{pdf_filename}"
    est_num = est.get("estimate_number") or str(est["id"])
    caption = f"Scope of Work - Estimate #{est_num}"

    existing = database.get_job_photos_by_job_week(est["job_id"], week_folder)
    already_exists = any(p.get("image_file") == rel_path for p in existing)

    if not already_exists:
        database.create_job_photo(
            job_id=est["job_id"],
            token_str=token_str,
            week_folder=week_folder,
            image_file=rel_path,
            thumb_file="",
            caption=caption,
        )

    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Admin: Client-facing estimate PDF
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/estimates/<int:estimate_id>/report/client-pdf")
def admin_estimate_client_pdf(estimate_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    est = database.get_estimate(estimate_id)
    if not est:
        return render_template("errors/404.html"), 404
    h._verify_token_access(est["token"])

    job = database.get_job(est["job_id"])
    job_name = job["job_name"] if job else "Unknown Job"
    items = database.get_estimate_items(estimate_id)
    token_data = database.get_token(est["token"])
    photos = database.get_all_job_photos_for_job(est["job_id"])

    if "photos" in request.args:
        try:
            selected_ids = {int(x) for x in request.args.get("photos", "").split(",") if x.strip()}
            photos = [p for p in photos if p["id"] in selected_ids]
        except ValueError:
            photos = []

    import pdf_generator
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_generator.generate_client_estimate_pdf(
            output_path=tmp.name,
            estimate=est,
            job=job,
            items=items,
            token_data=token_data,
            photos=photos,
        )
        _cust = re.sub(r"[^a-zA-Z0-9 \-]", "", (est.get("customer_name") or "").strip())[:25]
        _jname = re.sub(r"[^a-zA-Z0-9 \-]", "", job_name.strip())[:30]
        _date = (est.get("created_at") or "")[:10]
        _fname = "-".join(p for p in [_cust, _jname, _date] if p) + ".pdf"
        return send_file(
            tmp.name,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=_fname,
        )


# ---------------------------------------------------------------------------
# Admin: Job tasks management
# ---------------------------------------------------------------------------

@estimates_bp.route("/admin/job-tasks/<int:job_id>")
def admin_job_tasks(job_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return redirect(url_for("login"))

    job = database.get_job(job_id)
    if not job:
        return render_template("errors/404.html"), 404
    h._verify_token_access(job["token"])

    tasks = database.get_job_tasks(job_id)
    token_data = database.get_token(job["token"])

    _margin_target, _markup_required = _compute_finance_targets(job["token"], token_data)

    return render_template(
        "admin/estimate_detail.html",
        job=job,
        tasks=tasks,
        selected_token=token_data,
        estimate=None,
        photos=[],
        margin_target=_margin_target,
        markup_required=_markup_required,
    )


@estimates_bp.route("/admin/job-tasks/<int:job_id>/create", methods=["POST"])
def admin_create_job_task(job_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    job = database.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    h._verify_token_access(job["token"])

    name = request.form.get("name", "").strip()
    if not name:
        return jsonify({"error": "Task name required"}), 400

    task = database.create_job_task(job_id, job["token"], name, source="manual")
    return jsonify({"ok": True, "task": task})


@estimates_bp.route("/admin/job-tasks/<int:task_id>/toggle", methods=["POST"])
def admin_toggle_job_task(task_id):
    h = _helpers()
    if not h.current_user.is_authenticated:
        return jsonify({"error": "Unauthorized"}), 401

    task = database.get_job_task(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404
    h._verify_token_access(task["token"])

    task = database.toggle_job_task(task_id)
    return jsonify({"ok": True, "is_active": task["is_active"]})
